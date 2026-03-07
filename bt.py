"""
EQS V1.2 백테스트 (Edge Quant Signal)
=====================================================
EQS V1.0 → V1.1 변경사항:
  [1] ㊷ 코스피200 동적 유니버스 — 기존 고정 8종목에서 코스피200 전체로 확대
      - pykrx로 기준일 코스피200 구성종목 자동 조회
      - 시가총액+KRX업종 기반 클러스터 자동 분류 (대형가치/기술대형/금융/중소형성장)
      - 슬리피지 시총 기준 자동 결정 (대형0.003/중형0.005/소형0.008)
      - SECTOR_MAP 동적 구성
      - pykrx 미설치 또는 조회 실패 시 기존 8종목 폴백

EQS V1.0 대비 신규 추가 (V1.1):
  ㉝ RSI 다이버전스 감지 — 주가 신고가 + RSI 하락 시 Edge 감점
  ㉞ 섹터 로테이션 가중치 — 강세 섹터 소속 종목 Edge 보너스
  ㉟ 멀티 타임프레임 필터 — 주봉 추세 ≠ 일봉 신호 시 진입 차단
  ㊱ 일간 드로다운 서킷브레이커 — 당일 전체 평가손 -5% 시 신규진입 중단
  ㊲ 변동성 조절 포지션 사이징 — ATR 기반 종목당 리스크 균등화
  ㊳ 타임스탑 — N거래일 무변동 시 자동 청산
  ㊴ 섹터 집중도 한도 — 동일 섹터 3종목 이상 진입 차단
  ㊵ 주간 성과 분석 리포트 강화
  ㊶ 경제 이벤트 캘린더 — 주요 일정 전 Edge 커트라인 자동 상향

v26.0 대비 기존 추가:
  ㉚ ALT_FILTER_RATIO 국면 연동 동적 필터
     - 기존: 국면 무관 고정값 ALT_FILTER_RATIO = 3.5
     - 변경: 국면별 차등 적용
         BULL = 3.0  (추세장 → 정규 진입과 동일, 대안 배치 활성화)
         SIDE = 3.5  (횡보장 → 현재 기준 유지)
         BEAR = 4.0  (하락장 → 더 엄격, 현금 보유 우선)
     - 효과: "장세는 좋은데 시스템이 너무 매매를 안 한다" 문제 해결
     - 이중 허들(㉖필터 + ㉘기회비용)이 BULL에서 자동 완화
     - 엑셀 리포트에 국면별 필터 적용 이력 기록

  ㉛ 교체 마찰 비용 로직 (Hold Friction Cost)
     - 특정 종목 가산점 대신 "교체 자체의 비용"을 허들로 설정
     - 현재 보유 종목은 이미 진입 슬리피지를 지불한 상태
     - 교체 시 총 비용 = 현 종목 매도 슬리피지 × HOLD_FRICTION_MULT(1.5)
                        + 대안 종목 매수 슬리피지
     - 대안 Edge 우위가 교체 총비용 × 3.0 이상일 때만 교체 신호 발생
     - 효과: 모든 보유 종목에 동일 적용 → 백테스트 유효성 유지
             미미한 Edge 차이로 인한 잦은 교체 방지

v26.0 유지:
  ①~㉙ 전 기능 유지

필요 라이브러리:
    pip install yfinance pandas numpy scipy matplotlib requests
                beautifulsoup4 openpyxl schedule

실행 (즉시):
    python eqs_backtest_v1_1.py

실행 (스케줄러):
    python eqs_backtest_v1_1.py --scheduler
"""

import sys, os, json
from pathlib import Path as _Path

# ── .env 로드 (python-dotenv 있으면 자동, 없으면 수동 파싱) ──────────────
def _load_env():
    env_file = _Path(__file__).parent / ".env"
    if not env_file.exists():
        return
    try:
        from dotenv import load_dotenv
        load_dotenv(env_file)
    except ImportError:
        for line in env_file.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, _, v = line.partition("=")
                os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))

_load_env()
import pandas as pd
import numpy as np
from scipy import stats
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import time, warnings
from pathlib import Path
warnings.filterwarnings("ignore")
matplotlib.rcParams['font.family'] = 'DejaVu Sans'

try:
    import schedule; SCHEDULE_AVAILABLE = True
except ImportError:
    SCHEDULE_AVAILABLE = False

try:
    import yfinance as yf; YF_AVAILABLE = True
except ImportError:
    YF_AVAILABLE = False
    print("⚠️  yfinance 미설치 → 시뮬레이션 데이터로 실행합니다.")
    print("     python3 -m pip install yfinance")

# 하위 호환
FDR_AVAILABLE    = YF_AVAILABLE
PYKRX_AVAILABLE  = YF_AVAILABLE


# ══════════════════════════════════════════════════════
# 설정값
# ══════════════════════════════════════════════════════

START_DATE         = "20230101"
END_DATE           = "20260227"
MED_VOL_FALLBACK   = 0.016
AFTER_CLOSE_WEIGHT = 1.3
VOL_ANOMALY_MULT   = 3.0
VOL_ZONE_EXPAND    = 1.5

CORR_HIGH_THRESHOLD = 0.7
PORT_BETA_LIMIT     = 1.5

REGIME_MA_WINDOW  = 20
REGIME_EDGE_BULL  = 0.55
REGIME_EDGE_SIDE  = 0.60
REGIME_EDGE_BEAR  = 0.75

SELL_EDGE_THRESHOLD = 0.30

# ㉜ 주간 청산 원칙 (투자원칙 2)
MAX_HOLD_DAYS          = 5      # 최대 보유 거래일 (월~금 = 5일)
WEEKLY_CAPITAL_RESET   = True   # 매주 월요일 자본금 재계산 여부
FRIDAY_FORCE_EXIT      = True   # 금요일 강제 청산 여부
FRIDAY_HOLD_EDGE_THR   = 0.45   # 금요일 보유 유지 AI점수 기준 (리얼타임 동일)
# [Fix-1] RT is_friday_hold_ok와 완전 동일한 실효값 적용
# RT: hold_thr_eff = hold_thr - 0.03 / OPT: friday_hold_thr_eff = friday_hold_thr - 0.03
FRIDAY_HOLD_EDGE_THR_EFF = FRIDAY_HOLD_EDGE_THR - 0.03  # 실효 기준 0.42
CAPITAL_FLOOR_RATIO    = 0.70   # 자본 하방 보호 비율 (리얼타임 동일)


# ⑰ 클러스터별 슬리피지
SLIPPAGE_BY_CLUSTER = {
    "대형가치주":   0.003,
    "금융주":       0.003,
    "기술대형주":   0.003,   # [Issue-6 수정] RT LARGE_CAP 기반 0.003과 일치 (기존 "기타" 0.005 오적용 수정)
    "중소형성장주": 0.008,
    "기타":         0.005,
}
SLIPPAGE_TEST_LEVELS = [0.003, 0.005, 0.008]

# ⑭ 포지션 사이징
INITIAL_CAPITAL     = 10_000_000
KELLY_FRACTION      = 0.25
BEAR_SIZE_MULT      = 0.5
HIGH_CORR_SIZE_MULT = 0.7
MAX_POSITION_RATIO  = 0.30

# ⑱ 포트폴리오 총 비중 제한
EXPOSURE_CAP = {"BULL": 1.00, "SIDE": 0.70, "BEAR": 0.40}

# ㉒ ATR 동적 손절
STOP_LOSS_FALLBACK = -0.08
ATR_MULT_BY_CLUSTER = {
    "대형가치주":   1.2,
    "금융주":       1.2,
    "중소형성장주": 2.0,
    "기술대형주":   1.5,  # [H-2 수정] RT ATR_MULT_BY_CLUSTER_RT["기술대형주"]=1.5와 일치
    "기타":         1.5,
}
ATR_BEAR_MULT = 0.8
ATR_PERIOD    = 14
ATR_STOP_MIN  = 0.03
ATR_STOP_MAX  = 0.12

# ㉓ 트레일링 스탑
TRAIL_ACTIVATE_RET   = 0.07   # +7% 이상 시 트레일링 활성화
TRAIL_TIGHTEN_RET    = 0.25   # +25% 이상 시 간격 추가 축소
TRAIL_ATR_MULT = {            # 클러스터별 트레일링 ATR 계수
    "대형가치주":   1.5,
    "금융주":       1.5,
    "중소형성장주": 2.5,
    "기술대형주":   2.0,  # [H-2 수정] 기타와 동일 (RT TRAIL_ATR_MULT_DEFAULT=2.0과 일치)
    "기타":         2.0,
}
TRAIL_TIGHTEN_MULT   = 0.7    # +25% 이상 구간 추가 축소 배수
TAKE_PROFIT_FIXED    = 0.15   # 트레일링 미발동 구간 고정 익절

# ㉑ 슬리피지 진입 금지 필터
SLIPPAGE_FILTER_RATIO = 3.0

# ㉘ 기회비용 현실화 (한국 기준금리 3.5% ÷ 250 거래일)
ALT_OPPORTUNITY_COST = 0.00014  # 기존 0.0001 → 0.00014 (연 3.5%)

# ㉗ 트레일링 BEAR 국면 계수 (ATR_BEAR_MULT와 동일 값으로 일관성 유지)
TRAIL_BEAR_MULT      = 0.8

# ㉙ 대안 후보군 연산 효율
ALT_MAX_CANDIDATES   = 10   # Edge 상위 N개만 탐색
ALT_MAX_PENDING      = 5    # 동시 대기 최대 종목 수

# ㉔ 대안 종목 재배치
ALT_KELLY_MULT       = 0.7    # 대안 투입 Kelly 축소 계수
ALT_WAIT_DAYS        = 1      # 차단 후 재확인 대기 거래일 수

# ㉕㉖ 대안 배치 누적 슬리피지 + 엄격 필터
ALT_FILTER_RATIO     = 3.5    # 기본값 (하위호환용, 실제 적용은 ㉚ 국면 동적값)

# ㉚ 국면 연동 동적 ALT_FILTER_RATIO
ALT_FILTER_BY_REGIME = {
    "BULL": 3.0,   # 추세장: 정규 진입과 동일 → 대안 배치 활성화
    "SIDE": 3.5,   # 횡보장: 현재 기준 유지
    "BEAR": 4.0,   # 하락장: 더 엄격 → 현금 보유 우선
}

# ㉛ 교체 마찰 비용 (Hold Friction Cost)
HOLD_FRICTION_MULT   = 1.5    # 보유 종목 매도 슬리피지 가중 배수
HOLD_FRICTION_EDGE_GAP_MULT = 3.0  # 교체 비용 대비 Edge 우위 배수 기준
# ㉝ RSI 다이버전스 감지
RSI_DIVERGENCE_WINDOW  = 20     # RSI vs 주가 비교 윈도우
RSI_DIVERGENCE_PENALTY = 0.08   # 다이버전스 감지 시 Edge 감점

# ㉞ 섹터 로테이션 가중치
SECTOR_MOMENTUM_WINDOW = 20     # 섹터 모멘텀 산출 윈도우 (거래일)
SECTOR_MOMENTUM_BONUS  = 0.05   # 강세 섹터 Edge 보너스
SECTOR_MOMENTUM_PENALTY= 0.03   # 약세 섹터 Edge 감점
SECTOR_MAP = {}  # ㊷ build_kospi200_universe() 실행 후 아래에서 덮어씀

# ㉟ 멀티 타임프레임 필터
WEEKLY_MA_PERIOD       = 10     # 주봉 이동평균 기간 (10주 = 약 50거래일)
WEEKLY_TREND_REQUIRED  = True   # 주봉 추세 일치 여부 필터 사용

# ㊱ 일간 드로다운 서킷브레이커
DAILY_DRAWDOWN_LIMIT   = -0.05  # 당일 전체 평가손 한도
CIRCUIT_BREAKER_ENABLED= True   # 서킷브레이커 활성화

# ㊲ 변동성 조절 포지션 사이징
VOL_TARGET_DAILY       = 0.02   # 목표 일간 변동성 (2%)
VOL_TARGET_ENABLED     = True   # 변동성 타겟 사이징 활성화

# ㊳ 타임스탑
TIME_STOP_DAYS         = 15     # 최대 무변동 보유일
TIME_STOP_THRESHOLD    = 0.02   # 수익률 ±2% 이내면 무변동 판정
TIME_STOP_ENABLED      = True   # 타임스탑 활성화

# ㊴ 섹터 집중도 한도
SECTOR_MAX_POSITIONS   = 2      # 동일 섹터 최대 보유 종목 수
SECTOR_LIMIT_ENABLED   = True   # 섹터 집중도 제한 활성화

# ㊶ 경제 이벤트 캘린더
ECON_EVENT_EDGE_UPLIFT = 0.10   # 이벤트 당일 Edge 커트라인 상향
ECON_EVENTS_2025_2026  = [
    # (월, 일) — 주요 FOMC/CPI/고용 발표일 (근사값, 매년 업데이트 필요)
    (1,10),(1,15),(1,29),(2,7),(2,12),(2,26),(3,7),(3,12),(3,19),
    (4,4),(4,10),(4,30),(5,2),(5,13),(5,14),(6,6),(6,11),(6,18),
    (7,3),(7,11),(7,30),(8,1),(8,13),(8,27),(9,5),(9,10),(9,17),
    (10,3),(10,10),(10,29),(11,7),(11,12),(11,26),(12,5),(12,10),(12,17),
]


EMAIL_CONFIG = {
    "sender":   os.environ.get("EMAIL_SENDER",   "your_email@gmail.com"),
    "password": os.environ.get("EMAIL_PASSWORD", ""),
    "receiver": os.environ.get("EMAIL_RECEIVER", "receiver@email.com"),
    "smtp":     os.environ.get("EMAIL_SMTP",     "smtp.gmail.com"),
    "port":     int(os.environ.get("EMAIL_PORT", "587")),
}

# ══════════════════════════════════════════════════════
# ㊷ 코스피200 동적 유니버스 (EQS V1.1)
# ──────────────────────────────────────────────────────
# pykrx로 코스피200 종목을 실시간 조회하고
# 시가총액 + KRX 업종코드 기반으로 클러스터를 자동 분류한다.
#
# 클러스터 분류 규칙:
#   ① KRX 업종 → 금융 계열               → "금융주"
#   ② KRX 업종 → IT/반도체 + 시총 상위   → "기술대형주"
#   ② KRX 업종 → IT/반도체 + 시총 하위   → "중소형성장주"
#   ③ 시총 상위 (비금융/비IT)             → "대형가치주"
#   ④ 나머지                              → "중소형성장주"
#
# 슬리피지: 시총 기준 자동 결정
#   시총 10조 이상  → 0.003  (대형)
#   시총 1~10조     → 0.005  (중형)
#   시총 1조 미만   → 0.008  (소형)
# ══════════════════════════════════════════════════════

# ㊸ 연도별 코스피200 유니버스 (생존 편향 제거)
# YEARLY_UNIVERSE: {연도(int): set(종목코드)} — 진입 허용 여부 판단에 사용
# KOSPI200_REFERENCE_DATE: 클러스터 분류 기준 (가장 오래된 연도 기준일)
KOSPI200_REFERENCE_DATE = "20150102"   # 클러스터 분류 기준 (단일 시점)
YEARLY_UNIVERSE: dict = {}             # {year: set(code)} — build_yearly_kospi200()이 채움
LARGE_CAP_THRESHOLD_KRW = 10_000_000_000_000      # 10조 이상 → 대형
MID_CAP_THRESHOLD_KRW   =  1_000_000_000_000      #  1조 이상 → 중형

# KRX 업종명 → EQS 섹터명 매핑
_KRX_SECTOR_TO_EQS = {
    "금융": "금융", "은행": "금융", "증권": "금융", "보험": "금융", "기타금융": "금융",
    "전기전자": "IT전자", "반도체": "IT전자", "IT": "IT전자",
    "소프트웨어": "IT서비스", "통신": "IT서비스",
    "자동차": "자동차", "운수장비": "자동차",
    "철강금속": "소재", "화학": "소재", "에너지": "소재",
    "의약": "헬스케어", "의료정밀": "헬스케어",
}

# 클러스터별 파라미터 (tickers 리스트는 build_kospi200_universe()에서 채워짐)
CLUSTER_PARAMS = {
    "대형가치주": {
        "tickers": [],
        "W_MF":0.4, "W_TECH":0.3, "W_MOM":0.3,
        "T_DOF":7, "ASYM_BASE":1.2, "MF_CAP":0.08, "TECH_CAP":0.015,
    },
    "중소형성장주": {
        "tickers": [],
        "W_MF":0.3, "W_TECH":0.25, "W_MOM":0.45,
        "T_DOF":3, "ASYM_BASE":1.5, "MF_CAP":0.12, "TECH_CAP":0.03,
    },
    "금융주": {
        "tickers": [],
        "W_MF":0.45, "W_TECH":0.3, "W_MOM":0.25,
        "T_DOF":6, "ASYM_BASE":1.3, "MF_CAP":0.08, "TECH_CAP":0.012,
    },
    "기술대형주": {
        "tickers": [],
        "W_MF":0.35, "W_TECH":0.35, "W_MOM":0.30,
        "T_DOF":5, "ASYM_BASE":1.3, "MF_CAP":0.10, "TECH_CAP":0.020,
    },
}

def _classify_cluster(market_cap: int, krx_sector: str) -> str:
    """시가총액 + KRX 업종으로 EQS 클러스터 자동 결정."""
    eqs_sec  = _KRX_SECTOR_TO_EQS.get(krx_sector, "기타")
    is_large = market_cap >= LARGE_CAP_THRESHOLD_KRW
    if eqs_sec == "금융":
        return "금융주"
    if eqs_sec in ("IT전자", "IT서비스"):
        return "기술대형주" if is_large else "중소형성장주"
    if is_large:
        return "대형가치주"
    return "중소형성장주"

def _auto_slippage(market_cap: int) -> float:
    """시가총액 기준 슬리피지 자동 결정."""
    if market_cap >= LARGE_CAP_THRESHOLD_KRW:
        return 0.003
    if market_cap >= MID_CAP_THRESHOLD_KRW:
        return 0.005
    return 0.008

# fdr 코스피 종목 리스트 캐시
# ══════════════════════════════════════════════════════
# ㊷㊸ 코스피200 유니버스 (fdr 기반, StockListing 미사용)
# ──────────────────────────────────────────────────────
# fdr.StockListing("KOSPI")가 KRX JSON 파싱 오류로 불안정.
# → 코스피200 대표 종목 리스트 하드코딩 + fdr.DataReader로 OHLCV 수집
# 연도별 생존편향 제거: 상장일(IPO) 정보를 fdr.DataReader로 확인 불가하므로
#   실용적 대안으로 시총 규모별 클러스터를 사전 정의하고
#   YEARLY_UNIVERSE는 전 기간 동일 유니버스로 단순화
#   (생존편향 완전 제거보다 데이터 안정성 우선)
# ══════════════════════════════════════════════════════

# 코스피200 대표 종목 (2015~2026 기간 내 주요 상장 종목)
_KOSPI200_CODES = [
    # 반도체/IT
    "005930","000660","066570","009150","035420","251270","036570",
    # 자동차
    "005380","000270","012330","018880","010140",
    # 금융
    "105560","055550","086790","032830","316140","138930","024110",
    # 소재/에너지
    "005490","010130","011790","096770","010950","003670",
    # 헬스케어
    "068270","207940","128940","326030",
    # 소비재/유통
    "004370","097950","000810","051900","009830",
    # 기술/성장
    "035720","047050","259960","293490","373220","006400",
    # 중소형
    "031980","112610","051910","006260","042670","336260",
    # 건설/기계
    "000720","011200","009540","047810","034020",
    # 통신
    "017670","030200","032640",
]

# 종목코드 → (종목명, 클러스터, EQS섹터, 시총규모) 매핑
_TICKER_META = {
    # 기술대형주
    "005930": ("삼성전자",    "기술대형주", "IT전자",   "large"),
    "000660": ("SK하이닉스",  "기술대형주", "IT전자",   "large"),
    "066570": ("LG전자",      "기술대형주", "IT전자",   "large"),
    "009150": ("삼성전기",    "기술대형주", "IT전자",   "mid"),
    "035420": ("NAVER",       "기술대형주", "IT서비스", "large"),
    "251270": ("넷마블",      "중소형성장주","IT서비스","mid"),
    "036570": ("엔씨소프트",  "중소형성장주","IT서비스","mid"),
    "035720": ("카카오",      "기술대형주", "IT서비스", "large"),
    "047050": ("포스코인터내셔널","중소형성장주","소재", "mid"),
    "259960": ("크래프톤",    "중소형성장주","IT서비스","mid"),
    "293490": ("카카오뱅크",  "금융주",     "금융",     "mid"),
    "373220": ("LG에너지솔루션","기술대형주","소재",    "large"),
    "006400": ("삼성SDI",     "기술대형주", "소재",     "large"),
    # 자동차
    "005380": ("현대차",      "대형가치주", "자동차",   "large"),
    "000270": ("기아",        "대형가치주", "자동차",   "large"),
    "012330": ("현대모비스",  "대형가치주", "자동차",   "large"),
    "018880": ("한온시스템",  "중소형성장주","자동차",  "mid"),
    "010140": ("삼성중공업",  "중소형성장주","기타",    "mid"),
    # 금융
    "105560": ("KB금융",      "금융주",     "금융",     "large"),
    "055550": ("신한지주",    "금융주",     "금융",     "large"),
    "086790": ("하나금융지주","금융주",     "금융",     "large"),
    "032830": ("삼성생명",    "금융주",     "금융",     "large"),
    "316140": ("우리금융지주","금융주",     "금융",     "mid"),
    "138930": ("BNK금융지주", "금융주",     "금융",     "mid"),
    "024110": ("기업은행",    "금융주",     "금융",     "mid"),
    # 소재/에너지
    "005490": ("POSCO홀딩스", "기술대형주", "소재",     "large"),
    "010130": ("고려아연",    "대형가치주", "소재",     "mid"),
    "011790": ("SKC",         "중소형성장주","소재",    "mid"),
    "096770": ("SK이노베이션","대형가치주", "소재",     "large"),
    "010950": ("S-Oil",       "대형가치주", "소재",     "mid"),
    "003670": ("포스코퓨처엠","중소형성장주","소재",    "mid"),
    # 헬스케어
    "068270": ("셀트리온",    "중소형성장주","헬스케어","large"),
    "207940": ("삼성바이오로직스","중소형성장주","헬스케어","large"),
    "128940": ("한미약품",    "중소형성장주","헬스케어","mid"),
    "326030": ("SK바이오팜",  "중소형성장주","헬스케어","mid"),
    # 소비재
    "004370": ("농심",        "대형가치주", "기타",     "mid"),
    "097950": ("CJ제일제당",  "대형가치주", "기타",     "mid"),
    "000810": ("삼성화재",    "금융주",     "금융",     "large"),
    "051900": ("LG생활건강",  "대형가치주", "기타",     "large"),
    "009830": ("한화솔루션",  "중소형성장주","소재",    "mid"),
    # 중소형성장주
    "031980": ("피에스케이홀딩스","중소형성장주","IT전자","small"),
    "112610": ("씨에스윈드",  "중소형성장주","기타",    "small"),
    "051910": ("LG화학",      "대형가치주", "소재",     "large"),
    "006260": ("LS",          "대형가치주", "소재",     "mid"),
    "042670": ("HD현대인프라코어","중소형성장주","기타", "mid"),
    "336260": ("두산퓨얼셀",  "중소형성장주","기타",    "small"),
    # 건설/기계
    "000720": ("현대건설",    "대형가치주", "기타",     "mid"),
    "011200": ("HMM",         "대형가치주", "기타",     "mid"),
    "009540": ("HD한국조선해양","대형가치주","기타",    "mid"),
    "047810": ("한국항공우주","중소형성장주","기타",    "mid"),
    "034020": ("두산에너빌리티","중소형성장주","기타",  "mid"),
    # 통신
    "017670": ("SK텔레콤",    "대형가치주", "IT서비스", "large"),
    "030200": ("KT",          "대형가치주", "IT서비스", "large"),
    "032640": ("LG유플러스",  "대형가치주", "IT서비스", "mid"),
}

_SIZE_TO_SLIPPAGE = {"large": 0.003, "mid": 0.005, "small": 0.008}

def build_yearly_kospi200():
    """
    ㊸ 연도별 코스피200 유니버스.
    fdr.StockListing 불안정으로 _TICKER_META 하드코딩 사용.
    전 기간 동일 유니버스 적용 (YEARLY_UNIVERSE 필터 비활성화).
    """
    global YEARLY_UNIVERSE
    # StockListing 불안정 → 유니버스 필터 비활성화, 전 기간 동일 종목 사용
    YEARLY_UNIVERSE = {}
    print("  ℹ️  연도별 유니버스: 하드코딩 코스피200 사용 (yfinance 기반)")

def is_in_yearly_universe(ticker: str, date) -> bool:
    """YEARLY_UNIVERSE가 비어있으면 필터 비활성화 (전 종목 허용)."""
    if not YEARLY_UNIVERSE:
        return True
    year = date.year if hasattr(date, "year") else int(str(date)[:4])
    universe = YEARLY_UNIVERSE.get(year)
    return True if not universe else ticker in universe


def build_kospi200_universe():
    """
    fdr 기반 코스피200 유니버스 구성.
    _TICKER_META 하드코딩으로 TICKERS/CLUSTER_PARAMS/SECTOR_MAP 구성.
    """
    for cp in CLUSTER_PARAMS.values():
        cp["tickers"] = []

    if not YF_AVAILABLE:
        print("  ⚠️  yfinance 미설치 → 기존 8종목 폴백")
        _fb = {
            "삼성전자":"005930","현대차":"005380","KB금융":"105560",
            "피에스케이홀딩스":"031980","SK하이닉스":"000660",
            "카카오":"035720","POSCO홀딩스":"005490","LG에너지솔루션":"373220",
        }
        for code, cluster in {
            "005930":"기술대형주","000660":"기술대형주","005490":"기술대형주","373220":"기술대형주",
            "005380":"대형가치주","105560":"금융주","031980":"중소형성장주","035720":"중소형성장주",
        }.items():
            CLUSTER_PARAMS[cluster]["tickers"].append(code)
        return _fb, {"반도체":["005930","000660"],"금융":["105560"],"자동차":["005380"]}

    print("  ▶ 코스피200 유니버스 구성 중 (하드코딩 메타 기반)...")

    tickers_dict = {}
    sector_map   = {}
    slip_totals  = {k: [] for k in CLUSTER_PARAMS}

    for code, (name, cluster, eqs_sec, size) in _TICKER_META.items():
        slip = _SIZE_TO_SLIPPAGE.get(size, 0.005)
        tickers_dict[name] = code
        CLUSTER_PARAMS[cluster]["tickers"].append(code)
        sector_map.setdefault(eqs_sec, []).append(code)
        slip_totals[cluster].append(slip)

    for cname, slips in slip_totals.items():
        if slips:
            SLIPPAGE_BY_CLUSTER[cname] = float(np.median(slips))

    print(f"  ✅ 코스피200 유니버스 구성 완료: {len(tickers_dict)}종목")
    for cname, cp in CLUSTER_PARAMS.items():
        print(f"     {cname}: {len(cp['tickers'])}종목")

    return tickers_dict, sector_map

build_yearly_kospi200()                          # ㊸ 연도별 유니버스 먼저 빌드
TICKERS, _built_sector_map = build_kospi200_universe()
SECTOR_MAP.update(_built_sector_map)  # ㊷ 동적 섹터맵 반영
TOP_TICKERS_MEDVOL = list(TICKERS.values())[:10]

def get_cluster_params(ticker):
    for name, p in CLUSTER_PARAMS.items():
        if ticker in p["tickers"]: return name, p
    return "기타", {"W_MF":0.4,"W_TECH":0.3,"W_MOM":0.3,
                    "T_DOF":5,"ASYM_BASE":1.3,"MF_CAP":0.1,"TECH_CAP":0.02}



# ══════════════════════════════════════════════════════
# ㉒ ATR 계산 및 동적 손절선
# ══════════════════════════════════════════════════════

def calc_atr(df, period=ATR_PERIOD):
    if "고가" in df.columns and "저가" in df.columns:
        prev_close = df["종가"].shift(1)
        tr = pd.concat([
            df["고가"] - df["저가"],
            (df["고가"] - prev_close).abs(),
            (df["저가"] - prev_close).abs(),
        ], axis=1).max(axis=1)
        return tr.rolling(period).mean()
    else:
        sigma = df["종가"].pct_change().rolling(period).std()
        return sigma * df["종가"] * 1.5

def calc_dynamic_stop_loss(df, cluster_name, regime_ser):
    atr       = calc_atr(df)
    atr_ratio = atr / df["종가"]
    base_mult = ATR_MULT_BY_CLUSTER.get(cluster_name, 1.5)
    regime_al = regime_ser.reindex(df.index).ffill().fillna("SIDE")
    bear_adj  = regime_al.map({"BEAR": ATR_BEAR_MULT, "BULL": 1.0, "SIDE": 1.0})
    dynamic_sl = -(atr_ratio * base_mult * bear_adj)
    return dynamic_sl.clip(lower=-ATR_STOP_MAX, upper=-ATR_STOP_MIN)


# ══════════════════════════════════════════════════════
# ㉓ 트레일링 스탑 상태 관리
# ══════════════════════════════════════════════════════

class TrailingStopState:
    """
    ㉓ 보유 포지션별 트레일링 스탑 상태 추적
    - 수익 +7% 이상 → 트레일링 활성화
    - 고점(peak) 갱신마다 트레일 가격 자동 상향
    - +25% 이상 구간 → ATR 간격 추가 축소
    """
    def __init__(self):
        self.active      = False
        self.peak_price  = 0.0
        self.trail_price = 0.0   # 이 가격 하회 시 청산

    def update(self, curr_price: float, buy_price: float,
               atr_ratio: float, cluster_name: str,
               regime: str = "SIDE") -> tuple:
        """
        ㉗ regime 파라미터 추가 → BEAR 국면 트레일링 간격 0.8× 축소
        매일 호출. 트레일링 스탑 발동 여부 반환
        반환: (exit_flag, trigger_reason)
        """
        ret = (curr_price - buy_price) / buy_price

        if not self.active and ret >= TRAIL_ACTIVATE_RET:
            self.active     = True
            self.peak_price = curr_price

        if not self.active:
            return False, None

        if curr_price > self.peak_price:
            self.peak_price = curr_price

        # ATR 간격 계수 결정
        peak_ret   = (self.peak_price - buy_price) / buy_price
        base_mult  = TRAIL_ATR_MULT.get(cluster_name, 2.0)
        if peak_ret >= TRAIL_TIGHTEN_RET:
            base_mult *= TRAIL_TIGHTEN_MULT

        # ㉗ BEAR 국면 추가 축소 — 하락장에서 수익 빨리 확보
        if regime == "BEAR":
            base_mult *= TRAIL_BEAR_MULT

        trail_gap        = self.peak_price * atr_ratio * base_mult
        self.trail_price = self.peak_price - trail_gap

        if curr_price <= self.trail_price:
            bear_note = " [BEAR국면축소]" if regime == "BEAR" else ""
            return True, (f"트레일링스탑(고점{self.peak_price:,.0f}→"
                          f"현재{curr_price:,.0f}, 간격{trail_gap:,.0f}{bear_note})")
        return False, None

    def reset(self):
        self.active = False; self.peak_price = 0.0; self.trail_price = 0.0


# ══════════════════════════════════════════════════════
# ㉑ 슬리피지 진입 금지 필터
# ══════════════════════════════════════════════════════

def check_slippage_filter(curr_price, target_price, slippage,
                           filter_ratio=None):
    """㉑㉖ slippage filter. filter_ratio 미지정 시 SLIPPAGE_FILTER_RATIO 사용"""
    ratio = filter_ratio if filter_ratio is not None else SLIPPAGE_FILTER_RATIO
    if target_price <= curr_price:
        return False, "진입금지(Target≤현재가)"
    target_return  = (target_price - curr_price) / curr_price
    min_required   = slippage * 2 * ratio
    if target_return < min_required:
        return False, (f"진입금지(비용과다:{target_return:.2%}"
                       f"<필요{min_required:.2%})")
    return True, f"진입허용({target_return:.2%}≥{min_required:.2%})"


# ══════════════════════════════════════════════════════
# ㉕ 대안 배치 누적 슬리피지 체크
# ══════════════════════════════════════════════════════

def check_alt_cumulative_cost(curr_price: float,
                               target_price: float,
                               alt_slippage: float,
                               wait_days: int,
                               regime: str = "SIDE") -> tuple:
    """
    ㉕㉚ 대안 배치 누적 총 비용 계산 + 국면 연동 동적 필터
    총비용 = 기회비용 + 대안 왕복 슬리피지
    기회비용 = ALT_OPPORTUNITY_COST × wait_days

    ㉚ 국면별 필터 비율 자동 적용:
        BULL=3.0 / SIDE=3.5 / BEAR=4.0

    반환: (통과여부, 상세사유, 비용분석dict)
    """
    # ㉚ 국면 연동 필터 비율
    filter_ratio = ALT_FILTER_BY_REGIME.get(regime, ALT_FILTER_RATIO)

    if target_price <= curr_price:
        return False, "대안금지(Target≤현재가)", {}

    target_return    = (target_price - curr_price) / curr_price
    opportunity_cost = ALT_OPPORTUNITY_COST * wait_days
    roundtrip_slip   = alt_slippage * 2
    total_cost       = opportunity_cost + roundtrip_slip
    min_required     = total_cost * filter_ratio

    cost_detail = {
        "기대수익":     round(target_return,    4),
        "기회비용":     round(opportunity_cost, 4),
        "왕복슬리피지": round(roundtrip_slip,   4),
        "총비용":       round(total_cost,        4),
        "적용필터비율": filter_ratio,
        "필요최소수익": round(min_required,      4),
        "대기일":       wait_days,
        "국면":         regime,
    }

    if target_return < min_required:
        reason = (f"대안금지[{regime}국면필터{filter_ratio}×]"
                  f"(기대{target_return:.2%}"
                  f"<총비용기준{min_required:.2%}"
                  f"[기회{opportunity_cost:.3%}+슬리피지{roundtrip_slip:.2%}])")
        return False, reason, cost_detail

    reason = (f"대안허용[{regime}국면필터{filter_ratio}×]"
              f"(기대{target_return:.2%}≥{min_required:.2%})")
    return True, reason, cost_detail


# ══════════════════════════════════════════════════════
# ⑰ 슬리피지 유틸
# ══════════════════════════════════════════════════════

def get_slippage(cluster_name):
    return SLIPPAGE_BY_CLUSTER.get(cluster_name, 0.005)

def apply_slippage(price, side, slippage):
    return price * (1 + slippage) if side == "buy" else price * (1 - slippage)


# ══════════════════════════════════════════════════════
# ㉛ 교체 마찰 비용 (Hold Friction Cost)
# ══════════════════════════════════════════════════════

def check_hold_friction(held_edge: float,
                         alt_edge: float,
                         held_slip: float,
                         alt_slip: float) -> tuple:
    """
    ㉛ 현재 보유 종목 → 대안 종목 교체 시 마찰 비용 체크
    보유 종목은 이미 진입 슬리피지를 지불한 상태이므로
    교체 총비용(매도+매수 슬리피지)을 감안해도 Edge 우위가 명확할 때만 교체

    교체 총비용 = 현 종목 매도 슬리피지 × HOLD_FRICTION_MULT
                + 대안 종목 매수 슬리피지
    교체 허용 조건: alt_edge - held_edge > 교체총비용 × HOLD_FRICTION_EDGE_GAP_MULT

    반환: (교체허용여부, 상세사유, 비용dict)
    """
    sell_cost  = held_slip * HOLD_FRICTION_MULT    # 매도 마찰 (1.5배 가중)
    buy_cost   = alt_slip                          # 매수 비용
    total_friction = sell_cost + buy_cost          # 교체 총 마찰

    edge_gap   = alt_edge - held_edge
    min_gap    = total_friction * HOLD_FRICTION_EDGE_GAP_MULT  # 마찰의 3배 이상 우위 필요

    detail = {
        "현보유Edge":   round(held_edge,       4),
        "대안Edge":     round(alt_edge,         4),
        "Edge우위":     round(edge_gap,          4),
        "매도마찰":     round(sell_cost,          4),
        "매수비용":     round(buy_cost,           4),
        "교체총마찰":   round(total_friction,     4),
        "필요Edge우위": round(min_gap,            4),
    }

    if edge_gap <= min_gap:
        reason = (f"교체보류(Edge우위{edge_gap:.3f}"
                  f"≤마찰기준{min_gap:.3f}"
                  f"[매도{sell_cost:.3f}+매수{buy_cost:.3f}]×{HOLD_FRICTION_EDGE_GAP_MULT})")
        return False, reason, detail

    reason = (f"교체허용(Edge우위{edge_gap:.3f}"
              f">{min_gap:.3f}"
              f"[마찰{total_friction:.3f}×{HOLD_FRICTION_EDGE_GAP_MULT}])")
    return True, reason, detail

# ══════════════════════════════════════════════════════
# ㉝ RSI 다이버전스 감지
# ══════════════════════════════════════════════════════

def calc_rsi_divergence(df, window=RSI_DIVERGENCE_WINDOW):
    """
    주가 신고가 + RSI 하락 = 베어리시 다이버전스 → Edge 감점
    반환: bool Series (True = 다이버전스 감지)
    """
    rsi = calc_rsi(df)
    price_high = df["종가"].rolling(window).max() == df["종가"]
    rsi_declining = rsi < rsi.shift(1).rolling(window).max()
    divergence = price_high & rsi_declining
    return divergence.fillna(False)


# ══════════════════════════════════════════════════════
# ㉞ 섹터 로테이션 가중치
# ══════════════════════════════════════════════════════

def get_sector_for_ticker(ticker):
    """종목코드 → 섹터명"""
    for sector, tickers in SECTOR_MAP.items():
        if ticker in tickers:
            return sector
    return "기타"

def calc_sector_momentum(all_res_data, idx, window=SECTOR_MOMENTUM_WINDOW):
    """
    각 섹터의 최근 window일 평균 수익률 계산
    반환: {섹터명: 수익률}
    """
    sector_rets = {}
    for sector, tickers in SECTOR_MAP.items():
        rets = []
        for r in all_res_data:
            if r["ticker"] in tickers and idx in r["res"].index:
                prices = r["res"]["종가"]
                loc = prices.index.get_loc(idx)
                if loc >= window:
                    ret = (prices.iloc[loc] - prices.iloc[loc-window]) / prices.iloc[loc-window]
                    rets.append(ret)
        sector_rets[sector] = float(np.mean(rets)) if rets else 0.0
    return sector_rets

def get_sector_edge_adj(ticker, sector_rets):
    """
    섹터 모멘텀 기반 Edge 보정값 반환
    상위 33% 섹터: +보너스, 하위 33%: -감점
    """
    sector = get_sector_for_ticker(ticker)
    if not sector_rets:
        return 0.0
    sorted_sectors = sorted(sector_rets.items(), key=lambda x: x[1], reverse=True)
    n = len(sorted_sectors)
    top_third = [s[0] for s in sorted_sectors[:max(1, n//3)]]
    bot_third = [s[0] for s in sorted_sectors[-max(1, n//3):]]
    if sector in top_third:
        return SECTOR_MOMENTUM_BONUS
    elif sector in bot_third:
        return -SECTOR_MOMENTUM_PENALTY
    return 0.0


# ══════════════════════════════════════════════════════
# ㉟ 멀티 타임프레임 필터
# ══════════════════════════════════════════════════════

def calc_weekly_trend(df, period=WEEKLY_MA_PERIOD):
    """
    주봉 MA 상승 여부 판단
    종가를 주단위로 리샘플링 → MA 계산 → 상승 추세면 True
    반환: bool (일봉 마지막 날짜 기준)
    """
    weekly = df["종가"].resample("W-FRI").last().dropna()
    if len(weekly) < period + 1:
        return True  # 데이터 부족 시 필터 통과 (보수적이지 않지만 진입 차단 방지)
    ma = weekly.rolling(period).mean()
    return float(weekly.iloc[-1]) > float(ma.iloc[-1])


# ══════════════════════════════════════════════════════
# ㊱ 일간 드로다운 서킷브레이커
# ══════════════════════════════════════════════════════

def check_daily_drawdown(positions, all_res_data, idx, day_start_values):
    """
    당일 전체 포트폴리오 평가손이 DAILY_DRAWDOWN_LIMIT 이하면 True (차단)
    day_start_values: {name: 장시작 시 평가금액}
    반환: (차단여부, 현재 드로다운 비율)
    """
    if not CIRCUIT_BREAKER_ENABLED:
        return False, 0.0
    start_total = sum(day_start_values.values()) if day_start_values else 0
    if start_total <= 0:
        return False, 0.0
    curr_total = 0.0
    for r in all_res_data:
        name = r["name"]
        pos = positions[name]
        if pos["amount"] > 0 and idx in r["res"].index:
            cp = r["res"]["종가"].get(idx, pos["buy_price"])
            curr_val = pos["amount"] * cp / pos["buy_price"]
            curr_total += curr_val
    if start_total > 0:
        dd = (curr_total - start_total) / start_total
        return dd <= DAILY_DRAWDOWN_LIMIT, dd
    return False, 0.0


# ══════════════════════════════════════════════════════
# ㊲ 변동성 조절 포지션 사이징
# ══════════════════════════════════════════════════════

def calc_vol_target_size(df, capital, kelly_size, period=14):
    """
    ATR 기반 변동성 타겟 사이징
    목표: 종목당 일간 손익 변동이 VOL_TARGET_DAILY × capital 이내
    반환: min(kelly_size, vol_target_size)
    """
    if not VOL_TARGET_ENABLED:
        return kelly_size
    atr = calc_atr(df, period)
    if isinstance(atr, pd.Series):
        atr_val = atr.iloc[-1] if len(atr) > 0 else 0
    else:
        atr_val = float(atr) if atr else 0
    price = df["종가"].iloc[-1] if len(df) > 0 else 0
    if atr_val <= 0 or price <= 0:
        return kelly_size
    # 종목당 목표 변동금액 = capital × VOL_TARGET_DAILY
    target_vol_amount = capital * VOL_TARGET_DAILY
    # 1주당 일간 변동금액 ≈ ATR
    # 투자금액 = target_vol_amount / (ATR/price) = target_vol_amount × price / ATR
    vol_target_size = target_vol_amount * price / atr_val
    vol_target_size = min(vol_target_size, capital * MAX_POSITION_RATIO)
    return min(kelly_size, vol_target_size)


# ══════════════════════════════════════════════════════
# ㊳ 타임스탑
# ══════════════════════════════════════════════════════

def check_time_stop(hold_days, ret):
    """
    N거래일 보유 후 수익률이 ±THRESHOLD 이내면 청산 신호
    반환: (청산여부, 사유)
    """
    if not TIME_STOP_ENABLED:
        return False, None
    if hold_days >= TIME_STOP_DAYS and abs(ret) <= TIME_STOP_THRESHOLD:
        return True, f"타임스탑({hold_days}일보유, 수익{ret:+.1%}≤±{TIME_STOP_THRESHOLD:.0%})"
    return False, None


# ══════════════════════════════════════════════════════
# ㊴ 섹터 집중도 한도
# ══════════════════════════════════════════════════════

def check_sector_concentration(ticker, positions, all_res_data):
    """
    동일 섹터에 이미 SECTOR_MAX_POSITIONS개 보유 시 진입 차단
    반환: (진입허용여부, 사유)
    """
    if not SECTOR_LIMIT_ENABLED:
        return True, "섹터제한미사용"
    target_sector = get_sector_for_ticker(ticker)
    if target_sector == "기타":
        return True, "기타섹터(제한없음)"
    count = 0
    for r in all_res_data:
        if r["ticker"] == ticker:
            continue
        pos = positions.get(r["name"], {})
        if pos.get("amount", 0) > 0:
            if get_sector_for_ticker(r["ticker"]) == target_sector:
                count += 1
    if count >= SECTOR_MAX_POSITIONS:
        return False, f"섹터집중차단({target_sector}:{count}종목보유≥{SECTOR_MAX_POSITIONS})"
    return True, f"섹터허용({target_sector}:{count}/{SECTOR_MAX_POSITIONS})"


# ══════════════════════════════════════════════════════
# ㊶ 경제 이벤트 캘린더
# ══════════════════════════════════════════════════════

def is_econ_event_day(idx):
    """주요 경제 이벤트 당일 여부"""
    try:
        ts = pd.Timestamp(idx)
        return (ts.month, ts.day) in ECON_EVENTS_2025_2026
    except:
        return False

def is_econ_event_eve(idx):
    """주요 경제 이벤트 전날 여부"""
    try:
        ts = pd.Timestamp(idx)
        tomorrow = ts + pd.Timedelta(days=1)
        return (tomorrow.month, tomorrow.day) in ECON_EVENTS_2025_2026
    except:
        return False




# ══════════════════════════════════════════════════════
# ⑱ Global Exposure Cap 관리자
# ══════════════════════════════════════════════════════

class PortfolioExposureManager:
    def __init__(self, initial_capital):
        self.total_capital   = initial_capital
        self.invested_amount = 0.0
        self.current_regime  = "SIDE"
        self.prev_regime     = "SIDE"
        self.rebalance_log   = []

    @property
    def cash(self):
        return self.total_capital - self.invested_amount

    @property
    def exposure_ratio(self):
        return self.invested_amount / self.total_capital if self.total_capital > 0 else 0

    def get_cap(self, regime=None):
        return EXPOSURE_CAP.get(regime or self.current_regime, 0.70)

    def can_buy(self, invest_amount, regime):
        cap      = self.get_cap(regime)
        headroom = max(0, self.total_capital * cap - self.invested_amount)
        if headroom <= 0:
            return False, 0, f"비중한도초과({self.exposure_ratio:.1%}≥{cap:.0%})"
        allowed = min(invest_amount, headroom)
        reason  = (f"비중축소({invest_amount:,.0f}→{allowed:,.0f})"
                   if allowed < invest_amount else "정상")
        return True, allowed, reason

    def on_buy(self, amount):
        self.invested_amount += amount
        self.total_capital   -= amount

    def on_sell(self, amount, pnl):
        self.invested_amount -= amount
        self.total_capital   += amount + pnl

    def weekly_reset(self, new_capital: float, carry_invested: float = 0.0):
        """㉜ 매주 월요일: 총 자산(현금+평가금)을 새 자본금으로 재설정
        carry_invested: 이월 포지션 원투자금 합계.
          이월 없을 때 = 0 (기존 동작 유지)
          이월 있을 때 = 해당 금액을 invested_amount에 재등록하여
                        free_cash 과다 계산(이중투자) 방지.
        """
        self.total_capital   = new_capital
        self.invested_amount = carry_invested  # 이월 포지션 원투자금 재등록

    def update_regime(self, new_regime, idx, positions):

        forced = []
        if new_regime == self.current_regime: return forced
        old_regime = self.current_regime
        self.prev_regime = old_regime
        self.current_regime = new_regime
        new_cap = self.get_cap(new_regime)
        excess  = self.invested_amount - self.total_capital * new_cap
        if excess > 0:
            for name, pos in positions.items():
                if pos["amount"] <= 0: continue
                ratio = min(1.0, excess / self.invested_amount)
                forced.append({"name": name,
                    "reason": f"국면전환({old_regime}→{new_regime}) 리밸런싱",
                    "reduce": pos["amount"] * ratio, "ratio": ratio})
            self.rebalance_log.append({
                "날짜": str(idx)[:10], "전환": f"{old_regime}→{new_regime}",
                "변경한도": f"{self.get_cap(old_regime):.0%}→{new_cap:.0%}",
                "초과금액": round(excess, 0), "청산종목수": len(forced),
            })
        return forced


# ══════════════════════════════════════════════════════
# ⑪ 시장 국면
# ══════════════════════════════════════════════════════

def calc_market_regime(mkt, window=REGIME_MA_WINDOW):
    ma     = mkt["종가"].rolling(window).mean()
    regime = pd.Series("SIDE", index=mkt.index)
    regime[mkt["종가"] > ma * 1.02] = "BULL"
    regime[mkt["종가"] < ma * 0.98] = "BEAR"
    return regime

def get_regime_threshold(regime):
    return {"BULL": REGIME_EDGE_BULL, "SIDE": REGIME_EDGE_SIDE,
            "BEAR": REGIME_EDGE_BEAR}.get(regime, REGIME_EDGE_SIDE)


# ══════════════════════════════════════════════════════
# ① ASYM_K 동적화
# ══════════════════════════════════════════════════════

def calc_dynamic_asym_k(mkt_vol, asym_base=1.3):
    q75 = mkt_vol.rolling(60).quantile(0.75)
    q25 = mkt_vol.rolling(60).quantile(0.25)
    ak  = pd.Series(asym_base, index=mkt_vol.index)
    ak  = ak.where(mkt_vol < q75, asym_base * 1.5)
    ak  = ak.where(mkt_vol > q25, asym_base * 0.8)
    return ak.clip(1.0, 3.0)


# ══════════════════════════════════════════════════════
# ③⑤ KIND API + Lag
# ══════════════════════════════════════════════════════

POS_KW = ["수주","공급","양산","상향","돌파","흑자","독점","MOU","증설","호실적",
          "계약","신제품","특허","수출","성장"]
NEG_KW = ["지연","하회","분쟁","소송","매각","유상증자","과열","해지","이탈",
          "리스크","적자","취소","리콜","조사","제재"]
NEU_KW = ["보합","공개","일정","준비","검토","발표예정","유지","관망"]

def score_items(items):
    if not items: return 0.5
    p = n = u = 0.0
    for title, w in items:
        for kw in POS_KW:
            if kw in title: p += w
        for kw in NEG_KW:
            if kw in title: n += w
        for kw in NEU_KW:
            if kw in title: u += w * 0.5
    d = p + n + u
    return float(np.clip((((p - n) / d) + 1) / 2, 0.0, 1.0)) if d else 0.5

def fetch_kind(corp_code, date_str):
    try:
        resp = requests.get(
            "https://kind.krx.co.kr/disclosure/todaydisclosure.do",
            params={"method":"searchTodayDisclosureMain","currentPage":1,"maxResults":20,
                    "marketType":"kospi","searchFilter":"T","lstCrtDt":date_str,
                    "isuCd":corp_code,"typeCode":""},timeout=5)
        soup = BeautifulSoup(resp.text,"html.parser")
        return [{"title": r.select_one("td.disclosure-title").get_text(strip=True),
                 "time":  r.select_one("td.time").get_text(strip=True)
                          if r.select_one("td.time") else "09:00"}
                for r in soup.select("tr") if r.select_one("td.disclosure-title")]
    except: return []

def build_kind_mom(ticker, dates, use_api=True):
    if not use_api: return None
    print(f"    KIND API ({ticker})... ", end="", flush=True)
    d0, d1 = {}, {}
    for i, dt in enumerate(dates[-20:]):
        discs = fetch_kind(ticker, dt.strftime("%Y%m%d"))
        i0, i1 = [], []
        for disc in discs:
            try:
                h, m = int(disc["time"][:2]), int(disc["time"][3:5])
                after = h > 15 or (h == 15 and m >= 30)
            except: after = False
            (i1 if after else i0).append(
                (disc["title"], AFTER_CLOSE_WEIGHT if after else 1.0))
        d0[dt] = score_items(i0)
        if i + 1 < len(dates[-20:]):
            nxt = dates[-20:][i+1]
            d1[nxt] = (d1.get(nxt, 0.5) + score_items(i1)) / 2
        time.sleep(0.3)
    full = pd.Series(0.5, index=dates)
    for dt in dates[-20:]:
        full[dt] = d0.get(dt, 0.5) * 0.6 + d1.get(dt, 0.5) * 0.4
    print("완료"); return full


# ══════════════════════════════════════════════════════
# 기본 변수 산출
# ══════════════════════════════════════════════════════

def calc_mf(df, cap=0.1):
    return (df["foreign_net"].rolling(5).sum() /
            df["거래량"].rolling(20).mean()).fillna(0).clip(0, cap) / cap

def calc_tech(df, cap=0.02):
    ma20 = df["종가"].rolling(20).mean()
    return ((df["종가"] - ma20) / ma20).fillna(0).clip(0, cap) / cap

def calc_rsi(df, p=14):
    d = df["종가"].diff()
    g = d.clip(lower=0).rolling(p).mean()
    l = (-d.clip(upper=0)).rolling(p).mean()
    return 100 - (100 / (1 + g / l.replace(0, np.nan)))

def calc_mom_rsi(df, w=60):
    return calc_rsi(df).rolling(w).rank(pct=True).fillna(0.5)

def calc_sigma(df):
    ret = df["종가"].pct_change()
    return pd.concat([ret.rolling(5).std(), ret.rolling(20).std()], axis=1).max(axis=1)

def calc_beta_eff(df, mkt):
    rs = df["종가"].pct_change(); rm = mkt["종가"].pct_change()
    b20 = (rs.rolling(20).cov(rm) / rm.rolling(20).var()).fillna(1.0)
    b52 = (rs.rolling(252).cov(rm) / rm.rolling(252).var()).fillna(1.0)
    return b52.fillna(b20) * 0.3 + b20.fillna(b52) * 0.7

def calc_med_vol(mkt, w=5):
    if YF_AVAILABLE:
        vd = {}
        sd = pd.Timestamp(START_DATE).strftime("%Y-%m-%d")
        ed = pd.Timestamp(END_DATE).strftime("%Y-%m-%d")
        for tk in TOP_TICKERS_MEDVOL[:5]:   # 5개로 축소 (속도)
            try:
                tmp = yf.download(f"{tk}.KS", start=sd, end=ed,
                                  auto_adjust=True, progress=False)
                if tmp is None or len(tmp) < 10: continue
                tmp = _yf_parse(tmp)
                tmp.index = pd.to_datetime(tmp.index)
                if "종가" in tmp.columns:
                    vd[tk] = tmp["종가"].pct_change().abs()
            except: continue
        if vd:
            return (pd.DataFrame(vd).rolling(w).median().median(axis=1)
                    .reindex(mkt.index).ffill().fillna(MED_VOL_FALLBACK))
    return mkt["종가"].pct_change().abs().rolling(w).median().fillna(MED_VOL_FALLBACK)

def calc_ensemble_weights(df, mf, tech, mom, nc, window=20):
    actual_up = (nc > df["종가"]).astype(float)
    mf_acc   = (((mf   > 0.5).astype(float)) == actual_up).rolling(window).mean()
    tech_acc = (((tech > 0.5).astype(float)) == actual_up).rolling(window).mean()
    mom_acc  = (((mom  > 0.5).astype(float)) == actual_up).rolling(window).mean()
    raw = pd.concat([mf_acc, tech_acc, mom_acc], axis=1).fillna(1/3).clip(lower=0.1)
    total = raw.sum(axis=1)
    return raw.iloc[:,0]/total, raw.iloc[:,1]/total, raw.iloc[:,2]/total

def calc_vol_anomaly(df, mult=VOL_ANOMALY_MULT, window=20):
    avg_vol    = df["거래량"].rolling(window).mean()
    is_anomaly = df["거래량"] > avg_vol * mult
    return (is_anomaly,
            is_anomaly.map({True: VOL_ZONE_EXPAND, False: 1.0}),
            avg_vol, df["거래량"] / avg_vol)

def calc_targets(df, edge, sigma_sel, med_vol, beta_eff, asym_k_ser, ci_t, vol_zone_mult):
    cp     = df["종가"]; bias = (edge - 0.5) * 2
    sig_adj = pd.concat([sigma_sel, med_vol * bias.abs()], axis=1).max(axis=1)
    tp     = cp * (1 + sig_adj * beta_eff * bias)
    ci     = cp * sigma_sel * beta_eff * ci_t * vol_zone_mult
    d_mult = 1 + (asym_k_ser - 1) * (1 - edge)
    return tp, tp + ci, tp - ci * d_mult, bias, sig_adj, d_mult


# ══════════════════════════════════════════════════════
# ⑭ Kelly + ⑯ 진입이유
# ══════════════════════════════════════════════════════

def calc_kelly_position(win_rate, avg_win, avg_loss, regime, is_high_corr,
                        capital, kelly_mult=1.0):
    """kelly_mult: ㉔ 대안 종목 투입 시 ALT_KELLY_MULT 적용"""
    if avg_loss == 0 or win_rate == 0:
        return {"ratio": 0.05, "amount": capital * 0.05,
                "full_kelly": 0.05, "reason": "데이터부족→최소5%"}
    b = abs(avg_win / avg_loss); p = win_rate; q = 1 - p
    full_kelly = max(0.01, min((p * b - q) / b, 1.0))
    # [BUG-7 수정] RT calc_kelly_amount와 Kelly 배율 통일 (1/2 켈리)
    #       RT: full_kelly * 0.5 = 1/2 켈리 → BT 대비 포지션 2배 → 검증 불일치
    # 수정: BT도 1/2 켈리 적용, KELLY_FRACTION을 상한(cap)으로 유지
    #       OPT 최적화 KELLY_MAX_FRACTION이 BT KELLY_FRACTION으로 패치되므로
    #       cap = KELLY_FRACTION → BT/RT 포지션 크기 실질 동일
    kelly = full_kelly * 0.5 * kelly_mult
    kelly = min(kelly, KELLY_FRACTION)   # OPT 최적값 상한 적용
    if regime == "BEAR":
        kelly *= BEAR_SIZE_MULT; reason = f"BEAR→×{BEAR_SIZE_MULT}"
    elif regime == "BULL":
        reason = "BULL→기본Kelly"
    else:
        reason = "SIDE→기본Kelly"
    if is_high_corr:
        kelly *= HIGH_CORR_SIZE_MULT; reason += f"+고상관×{HIGH_CORR_SIZE_MULT}"
    if kelly_mult < 1.0:
        reason += f"+대안종목×{kelly_mult}"
    kelly = min(kelly, MAX_POSITION_RATIO)
    return {"ratio": round(kelly, 4), "amount": round(capital * kelly, 0),
            "full_kelly": round(full_kelly, 4), "reason": reason}

def analyze_entry_reason(row, cp):
    mf_c   = row["mf"]   * cp.get("W_MF",   0.4)
    tech_c = row["tech"] * cp.get("W_TECH",  0.3)
    mom_c  = row["mom"]  * cp.get("W_MOM",   0.3)
    total  = mf_c + tech_c + mom_c; reasons = []
    if total > 0:
        if mf_c   / total >= 0.40: reasons.append(f"수급강세({row['mf']:.2f})")
        if tech_c / total >= 0.35: reasons.append(f"기술적돌파({row['tech']:.2f})")
        if mom_c  / total >= 0.35: reasons.append(f"모멘텀강화({row['mom']:.2f})")
    if row.get("vol_anomaly", False):
        reasons.append(f"거래량급증({row.get('vol_ratio',0):.1f}x)")
    return " + ".join(reasons) if reasons else "복합신호"


# ══════════════════════════════════════════════════════
# ⑮㉒㉓ 통합 청산 판단
# ══════════════════════════════════════════════════════

def check_exit_trigger(curr_price, buy_price, regime, dynamic_sl=None,
                       trailing: TrailingStopState = None,
                       atr_ratio: float = 0.02,
                       cluster_name: str = "기타"):
    """
    ㉒ ATR 동적 손절 + ㉓㉗ 트레일링 스탑(BEAR 국면 반영) + 고정 익절 통합 판단
    우선순위: 트레일링스탑 > ATR손절 > 고정익절
    """
    ret = (curr_price - buy_price) / buy_price

    # ㉓㉗ 트레일링 스탑 (regime 전달)
    if trailing is not None:
        trail_exit, trail_reason = trailing.update(
            curr_price, buy_price, atr_ratio, cluster_name, regime)
        if trail_exit:
            return True, trail_reason, ret

    # ㉒ ATR 동적 손절
    sl = float(dynamic_sl) if dynamic_sl is not None \
         else STOP_LOSS_FALLBACK * (0.625 if regime == "BEAR" else 1.0)
    if ret <= sl:
        return True, f"ATR손절({ret:+.1%}≤{sl:.1%})", ret

    # 고정 익절 (트레일링 미발동 구간)
    if trailing is None or not trailing.active:
        if ret >= TAKE_PROFIT_FIXED:
            return True, f"고정익절({ret:+.1%}≥{TAKE_PROFIT_FIXED:.0%})", ret

    return False, None, ret


def is_friday(idx) -> bool:
    """㉜ 금요일 여부 (주간 강제 청산 판단용)
    [Issue-5 수정] fdr 이용 가능 시 실제 휴장일(공휴일) 제외
    이전: weekday()==4 단순 체크 → 한국 공휴일 금요일에도 강제청산 트리거
    수정: 공휴일 금요일은 실제로 거래 없음 → 청산 스킵
    """
    try:
        ts = pd.Timestamp(idx)
        if ts.weekday() != 4:
            return False
        # yfinance로 실제 거래일 여부 확인
        if YF_AVAILABLE:
            try:
                d_str = ts.strftime("%Y-%m-%d")
                chk = yf.download("005930.KS", start=d_str, end=d_str,
                                   progress=False, auto_adjust=True)
                return chk is not None and len(chk) > 0
            except:
                pass
        return True   # yfinance 실패 시 금요일로 간주 (보수적)
    except:
        return False


# ══════════════════════════════════════════════════════
# ⑥ 적응형 Walk-forward
# ══════════════════════════════════════════════════════

def adaptive_walk_forward(res, train_weeks=8, ci_t=1.1, asym_base=1.3,
                          hr_hold=0.70, hr_retune=0.60, roll_w=2):
    candidates = [(mf, tech, round(1 - mf - tech, 1))
                  for mf in [0.3, 0.4, 0.5] for tech in [0.2, 0.3, 0.4]
                  if 0.1 <= round(1 - mf - tech, 1) <= 0.6]
    res = res.copy(); res["week"] = res.index.to_period("W")
    weeks = res["week"].unique(); records = []; cur_w = (0.4, 0.3, 0.3); hr_hist = []
    for i, week in enumerate(weeks):
        wd = res[res["week"] == week]
        e  = wd["mf"]*cur_w[0] + wd["tech"]*cur_w[1] + wd["mom"]*cur_w[2]
        b  = (e - 0.5) * 2
        sa = pd.concat([wd["sigma_sel"], wd["med_vol"]*b.abs()], axis=1).max(axis=1)
        tp = wd["종가"] * (1 + sa * wd["beta"] * b)
        ci = wd["종가"] * wd["sigma_sel"] * wd["beta"] * ci_t
        dm = 1 + (asym_base - 1) * (1 - e)
        iz = (wd["nc"] >= tp - ci * dm) & (wd["nc"] <= tp + ci)
        thr = iz.mean() if len(iz) else 0.5; hr_hist.append(thr)
        records.append({"week": week, "W_MF": cur_w[0], "W_TECH": cur_w[1],
                         "W_MOM": cur_w[2], "hr": thr, "action": "초기화"})
        if i < train_weeks: continue
        roll_hr = np.mean(hr_hist[-roll_w:])
        if roll_hr >= hr_hold:
            records[-1]["action"] = f"유지({roll_hr:.1%}≥{hr_hold:.0%})"; continue
        tight = roll_hr >= hr_retune
        cands = [w for w in candidates
                 if abs(w[0]-cur_w[0]) <= (0.1 if tight else 0.2)
                 and abs(w[1]-cur_w[1]) <= (0.1 if tight else 0.2)] or [cur_w]
        train = res[res["week"].isin(weeks[i - train_weeks:i])]
        best_hr, best_w = 0, cur_w
        for wc in cands:
            e2  = train["mf"]*wc[0] + train["tech"]*wc[1] + train["mom"]*wc[2]
            b2  = (e2 - 0.5) * 2
            sa2 = pd.concat([train["sigma_sel"], train["med_vol"]*b2.abs()], axis=1).max(axis=1)
            tp2 = train["종가"] * (1 + sa2 * train["beta"] * b2)
            ci2 = train["종가"] * train["sigma_sel"] * train["beta"] * ci_t
            dm2 = 1 + (asym_base - 1) * (1 - e2)
            hz  = ((train["nc"] >= tp2 - ci2*dm2) & (train["nc"] <= tp2 + ci2)).mean()
            if hz > best_hr: best_hr, best_w = hz, wc
        changed = cur_w != best_w; cur_w = best_w
        records[-1].update({"W_MF": cur_w[0], "W_TECH": cur_w[1], "W_MOM": cur_w[2]})
        records[-1]["action"] = (
            f"{'소폭' if tight else '전체'}튜닝({roll_hr:.1%})→"
            f"{'변경' if changed else '유지'}({best_hr:.1%})")
    return pd.DataFrame(records).set_index("week")


# ══════════════════════════════════════════════════════
# ⑩ 포트폴리오 리스크
# ══════════════════════════════════════════════════════

def calc_portfolio_risk(all_results):
    ret_dict  = {r["name"]: r["res"]["종가"].pct_change() for r in all_results}
    beta_dict = {r["name"]: r["res"]["beta"].mean()       for r in all_results}
    ret_df    = pd.DataFrame(ret_dict).dropna()
    corr_mat  = ret_df.corr()
    port_beta = sum(beta_dict.values()) / len(all_results)
    high_corr = [(n1, n2, corr_mat.loc[n1, n2])
                 for i, n1 in enumerate(corr_mat.columns)
                 for n2 in corr_mat.columns[i+1:]
                 if abs(corr_mat.loc[n1, n2]) >= CORR_HIGH_THRESHOLD]
    return {"corr_matrix": corr_mat, "port_beta": port_beta,
            "beta_dict": beta_dict, "high_corr_pairs": high_corr,
            "is_high_corr": len(high_corr) > 0}


# ══════════════════════════════════════════════════════
# ⑳ 슬리피지 민감도 분석
# ══════════════════════════════════════════════════════

def calc_slippage_sensitivity(res, mkt, regime_ser, cluster_name, cp):
    results = {}
    for slip in SLIPPAGE_TEST_LEVELS:
        capital = INITIAL_CAPITAL
        position = buy_price = prev_price = pos_amount = 0
        regime = regime_ser.reindex(res.index).ffill().fillna("SIDE")
        for idx, row in res.iterrows():
            rg        = regime.get(idx, "SIDE")
            threshold = get_regime_threshold(rg)
            cp_now    = row["종가"]
            if position == 1:
                exit_flag, _, _ = check_exit_trigger(
                    cp_now, buy_price, rg, row.get("dynamic_sl", None))
                if exit_flag or row["edge"] < SELL_EDGE_THRESHOLD:
                    sell_p   = apply_slippage(cp_now, "sell", slip)
                    capital += pos_amount * (sell_p - buy_price) / buy_price + pos_amount
                    position = buy_price = prev_price = pos_amount = 0
                else:
                    prev_price = cp_now
            if position == 0 and row["edge"] >= threshold:
                invest     = capital * 0.05
                buy_price  = apply_slippage(cp_now, "buy", slip)
                prev_price = buy_price; pos_amount = invest
                capital   -= invest; position = 1
        results[slip] = (capital / INITIAL_CAPITAL) - 1
    bep   = None; slips = sorted(results.keys())
    for i in range(len(slips) - 1):
        r1, r2 = results[slips[i]], results[slips[i+1]]
        if r1 >= 0 >= r2:
            bep = slips[i] + (slips[i+1] - slips[i]) * r1 / (r1 - r2); break
    if bep is None and results[slips[-1]] >= 0: bep = slips[-1]
    results["bep"] = bep; return results


# ══════════════════════════════════════════════════════
# 포트폴리오 통합 시뮬레이션 (㉑㉒㉓㉔ 반영)
# ══════════════════════════════════════════════════════

def calc_portfolio_pnl(all_res_data, regime_ser, is_high_corr=False):
    mgr       = PortfolioExposureManager(INITIAL_CAPITAL)
    regime    = regime_ser.ffill().fillna("SIDE")
    all_dates = all_res_data[0]["res"].index

    # 포지션 상태 (㉓ 트레일링 스탑 상태 추가)
    positions = {r["name"]: {
        "amount": 0, "buy_price": 0, "prev_price": 0,
        "hold_days": 0, "entry_row": None, "entry_reason": "",
        "trailing": TrailingStopState(),   # ㉓
        "peak_ret": 0.0,                   # ㉓ 최고 수익률 추적
        "sector": get_sector_for_ticker(r.get("ticker","")),  # ㊴ 섹터
    } for r in all_res_data}

    trade_logs       = {r["name"]: [] for r in all_res_data}
    capital_hist     = []
    rebalance_events = []
    blocked_log      = []   # ㉑ 차단 로그
    alt_log          = []   # ㉔ 대안 배치 로그
    trail_log        = []   # ㉓ 트레일링 발동 로그
    friction_log     = []   # ㉛ 교체 마찰 보류 로그
    time_stop_log    = []   # ㊳ 타임스탑 로그
    circuit_log      = []   # ㊱ 서킷브레이커 로그
    sector_block_log = []   # ㊴ 섹터집중 차단 로그

    # ㉔ 대안 대기 큐: {name: 차단된 날짜 인덱스}
    alt_pending = {}   # {blocked_name: {"blocked_idx": idx, "kelly_mult": float}}

    # ㊱ 서킷브레이커 일간 시작 평가금액
    day_start_values = {}
    prev_day = None
    circuit_active = False

    # ㉞ 섹터 모멘텀 캐시
    _sector_rets = {}

    # [New-B] 익일 시가 진입 대기 큐 — Look-ahead Bias 방지
    pending_entries = {}  # {name: {"signal_date": idx, "signal_row": row, "entry_reason": str,
                          #         "kelly_info": dict, "allowed": float}}

    # [Fix-BUG-5] 더미 Kelly → trade_log 실거래 통계 주입
    #       OPT가 Kelly=0.33 추천해도 BT 검증은 다른 포지션 크기로 진행 → 검증 무의미
    # 수정: trade_log.json이 있으면 실거래 통계로 초기화 (파이프라인 일관성 확보)
    #       없으면 보수적 기본값 유지 (기존 동작과 동일)
    _bt_win_rate = 0.40
    _bt_avg_win  = TAKE_PROFIT_FIXED
    _bt_avg_loss = abs(STOP_LOSS_FALLBACK)
    try:
        _tlog_path = Path(__file__).parent / "trade_log.json"
        if _tlog_path.exists():
            _tlogs = json.loads(_tlog_path.read_text(encoding="utf-8"))
            _closed = [t for t in _tlogs
                       if t.get("exit_price", 0) > 0 and t.get("buy_price", 0) > 0
                       and not t.get("carry_over", False)]
            if len(_closed) >= 10:   # 최소 10건 이상일 때만 실거래 통계 적용
                _rets  = [(t["exit_price"] - t["buy_price"]) / t["buy_price"] for t in _closed]
                _wins  = [r for r in _rets if r > 0]
                _losses= [abs(r) for r in _rets if r < 0]
                _bt_win_rate = len(_wins) / len(_rets)
                _bt_avg_win  = float(np.mean(_wins))  if _wins   else TAKE_PROFIT_FIXED
                _bt_avg_loss = float(np.mean(_losses)) if _losses else abs(STOP_LOSS_FALLBACK)
    except Exception:
        pass   # 실패 시 기본값 사용

    dummy_win_rate = _bt_win_rate
    dummy_avg_win  = _bt_avg_win
    dummy_avg_loss = _bt_avg_loss

    # [Issue-1 수정] 주차 시작 자본 추적 — 하방 보호 기준으로 사용 (RT monday_reset과 동일 방식)
    week_start_capital = INITIAL_CAPITAL

    # 날짜 → 인덱스 매핑
    date_list = list(all_dates)

    for day_i, idx in enumerate(all_dates):
        curr_regime = regime.get(idx, "SIDE")

        # ㊱ 서킷브레이커: 일간 시작 평가금액 기록
        curr_day = idx.date() if hasattr(idx, 'date') else idx
        if curr_day != prev_day:
            day_start_values = {}
            for r_data in all_res_data:
                nm = r_data["name"]
                ps = positions[nm]
                if ps["amount"] > 0 and idx in r_data["res"].index:
                    cp_start = r_data["res"]["종가"].get(idx, ps["buy_price"])
                    day_start_values[nm] = ps["amount"] * cp_start / ps["buy_price"]
            prev_day = curr_day
            circuit_active = False  # 새 날 시작 시 해제

        # ㊱ 서킷브레이커 체크
        if CIRCUIT_BREAKER_ENABLED and not circuit_active:
            cb_triggered, cb_dd = check_daily_drawdown(
                positions, all_res_data, idx, day_start_values)
            if cb_triggered:
                circuit_active = True
                circuit_log.append({
                    "날짜": str(idx)[:10], "드로다운": f"{cb_dd:.1%}",
                    "조치": "당일 신규진입 중단"})

        # ㊶ 경제 이벤트 당일 Edge 커트라인 상향
        econ_uplift = ECON_EVENT_EDGE_UPLIFT if is_econ_event_day(idx) else 0.0

        # ㉞ 섹터 모멘텀 (5일마다 갱신으로 연산 효율화)
        if day_i % 5 == 0:
            _sector_rets = calc_sector_momentum(all_res_data, idx)
        # else: 이전 값 유지 (루프 바깥 {} 초기화)

        # ⑲ 국면 전환 리밸런싱
        forced = mgr.update_regime(curr_regime, idx, positions)
        for f in forced:
            name = f["name"]; pos = positions[name]
            if pos["amount"] <= 0: continue
            r_data = next((r for r in all_res_data if r["name"] == name), None)
            if r_data is None: continue
            slip    = get_slippage(r_data["cluster"])
            cp_now  = r_data["res"]["종가"].get(idx, pos["buy_price"])
            sell_p  = apply_slippage(cp_now, "sell", slip)
            reduce  = min(f["reduce"], pos["amount"])
            pnl_amt = reduce * (sell_p - pos["buy_price"]) / pos["buy_price"]
            mgr.on_sell(reduce, pnl_amt)
            pos["amount"] -= reduce
            pos["trailing"].reset()
            _log_trade(trade_logs[name], name, pos, idx, curr_regime,
                       f["reason"], cp_now, sell_p, reduce, pnl_amt, None)
            rebalance_events.append({"날짜": str(idx)[:10], "종목": name,
                                     "사유": f["reason"], "금액": round(reduce, 0)})

        for r_data in all_res_data:
            name    = r_data["name"]
            res     = r_data["res"]
            cp_map  = r_data["cp"]
            cluster = r_data["cluster"]
            slip    = get_slippage(cluster)
            pos     = positions[name]
            if idx not in res.index: continue
            row    = res.loc[idx]; cp_now = row["종가"]
            threshold = get_regime_threshold(curr_regime)

            # [New-B] 전일 신호 → 금일 시가 진입 처리 (pending_entries)
            if name in pending_entries and pos["amount"] == 0:
                pend = pending_entries.pop(name)
                # next_open: T일의 res["next_open"] = T+1일의 시가
                entry_open = row["종가"]   # 당일 종가를 폴백으로 사용
                if "next_open" in pend["signal_row"].index:
                    no = pend["signal_row"]["next_open"]
                    if pd.notna(no) and no > 0:
                        entry_open = float(no)
                buy_p = apply_slippage(entry_open, "buy", slip)
                allowed = pend["allowed"]
                ok2, allowed2, _ = mgr.can_buy(allowed, curr_regime)
                if ok2 and allowed2 > 1000:
                    mgr.on_buy(allowed2)
                    pos.update({"amount": allowed2, "buy_price": buy_p,
                                "prev_price": buy_p, "hold_days": 0,
                                "entry_row": idx, "entry_reason": pend["entry_reason"],
                                "peak_ret": 0.0})
                    pos["trailing"].reset()
                    trade_logs[name].append({
                        "종목": name, "진입일": str(idx)[:10], "청산일": None,
                        "보유기간": None, "국면": curr_regime,
                        "진입이유": pend["entry_reason"] + "[T+1시가]",
                        "청산이유": "보유중",
                        "진입가": round(buy_p, 0), "청산가": None,
                        "수익률": None, "투자금": round(allowed2, 0),
                        "손익금": None, "Kelly비율": round(allowed2/INITIAL_CAPITAL, 4),
                        "ATR손절선": None, "㉔대안": "정규진입", "승패": "-",
                    })

            # ㉒ ATR
            dynamic_sl = row.get("dynamic_sl", None)
            atr_ratio  = abs(float(dynamic_sl)) if dynamic_sl is not None else 0.02

            # ─ 보유 중 처리 ─
            if pos["amount"] > 0:
                pos["hold_days"] += 1
                curr_ret = (cp_now - pos["buy_price"]) / pos["buy_price"]
                pos["peak_ret"] = max(pos["peak_ret"], curr_ret)

                exit_flag, trigger, _ = check_exit_trigger(
                    cp_now, pos["buy_price"], curr_regime,
                    dynamic_sl, pos["trailing"], atr_ratio, cluster)

                # ㊳ 타임스탑 체크 (ATR/트레일링보다 후순위)
                if not exit_flag:
                    ts_exit, ts_reason = check_time_stop(pos["hold_days"], curr_ret)
                    if ts_exit:
                        exit_flag = True
                        trigger = ts_reason
                        time_stop_log.append({
                            "날짜": str(idx)[:10], "종목": name,
                            "보유일": pos["hold_days"],
                            "수익률": f"{curr_ret:+.1%}"})

                # ─ Edge 하락 체크 + ㉛ 교체 마찰 비용 검증 ─
                edge_exit = row["edge"] < SELL_EDGE_THRESHOLD
                if edge_exit and not exit_flag:
                    # 대안 종목 중 가장 높은 Edge 찾기
                    best_alt = max(
                        (r for r in all_res_data
                         if r["name"] != name
                         and positions[r["name"]]["amount"] == 0
                         and idx in r["res"].index),
                        key=lambda r: r["res"].loc[idx, "edge"] if idx in r["res"].index else 0,
                        default=None)

                    if best_alt is not None and idx in best_alt["res"].index:
                        alt_edge = best_alt["res"].loc[idx, "edge"]
                        alt_slip = get_slippage(best_alt["cluster"])
                        friction_ok, friction_reason, friction_detail = check_hold_friction(
                            row["edge"], alt_edge, slip, alt_slip)
                        if not friction_ok:
                            # 대안이 압도적이지 않음 → 교체 보류, 청산도 보류
                            edge_exit = False
                            friction_log.append({
                                "날짜":       str(idx)[:10],
                                "보유종목":   name,
                                "대안종목":   best_alt["name"],
                                "결과":       "교체보류(마찰비용)",
                                "보유Edge":   round(row["edge"], 3),
                                "대안Edge":   round(alt_edge, 3),
                                "Edge우위":   round(friction_detail.get("Edge우위", 0), 3),
                                "필요우위":   round(friction_detail.get("필요Edge우위", 0), 3),
                                "교체총마찰": round(friction_detail.get("교체총마찰", 0), 4),
                                "사유":       friction_reason,
                            })
                        else:
                            friction_log.append({
                                "날짜":       str(idx)[:10],
                                "보유종목":   name,
                                "대안종목":   best_alt["name"],
                                "결과":       "교체허용",
                                "보유Edge":   round(row["edge"], 3),
                                "대안Edge":   round(alt_edge, 3),
                                "Edge우위":   round(friction_detail.get("Edge우위", 0), 3),
                                "필요우위":   round(friction_detail.get("필요Edge우위", 0), 3),
                                "교체총마찰": round(friction_detail.get("교체총마찰", 0), 4),
                                "사유":       friction_reason,
                            })
                    else:
                        # 대안 없음 → 교체 비교 불가, 원래대로 청산
                        pass

                sell_now = exit_flag or edge_exit
                if not sell_now:
                    trigger = None

                if sell_now:
                    sell_p  = apply_slippage(cp_now, "sell", slip)
                    trade_r = (sell_p - pos["buy_price"]) / pos["buy_price"]
                    pnl_amt = pos["amount"] * trade_r
                    mgr.on_sell(pos["amount"], pnl_amt)
                    # ㉓ 트레일링 발동 기록
                    if trigger and "트레일링" in trigger:
                        trail_log.append({
                            "날짜": str(idx)[:10], "종목": name,
                            "최고수익률": f"{pos['peak_ret']:+.1%}",
                            "청산수익률": f"{trade_r:+.1%}",
                            "트레일수익보전": f"{trade_r - 0:.1%}",
                        })
                    exit_label = trigger if trigger else f"Edge하락({row['edge']:.2f})"
                    _log_trade(trade_logs[name], name, pos, idx, curr_regime,
                               exit_label, cp_now, sell_p, pos["amount"], pnl_amt, dynamic_sl)
                    pos.update({"amount": 0, "buy_price": 0, "prev_price": 0,
                                "hold_days": 0, "entry_row": None, "entry_reason": "",
                                "peak_ret": 0.0})
                    pos["trailing"].reset()
                else:
                    pos["prev_price"] = cp_now

            # ─ 신규 매수 ─
            # ㉝ RSI 다이버전스 감점
            adj_edge = row["edge"]
            if row.get("rsi_divergence", False):
                adj_edge -= RSI_DIVERGENCE_PENALTY
            # ㉞ 섹터 로테이션 보너스/감점
            adj_edge += get_sector_edge_adj(r_data["ticker"], _sector_rets)
            # ㊶ 경제 이벤트 당일 커트라인 상향
            adj_threshold = threshold + econ_uplift

            if pos["amount"] == 0 and adj_edge >= adj_threshold:
                # ㊸ 연도별 코스피200 유니버스 필터 (생존 편향 제거)
                if YEARLY_UNIVERSE and not is_in_yearly_universe(r_data["ticker"], idx):
                    blocked_log.append({
                        "날짜": str(idx)[:10], "종목": name, "국면": curr_regime,
                        "Edge": round(adj_edge, 3),
                        "현재가": round(cp_now, 0), "Target_P": 0,
                        "슬리피지": slip, "차단이유": f"㊸연도별유니버스미포함({idx.year})"})
                    continue
                # ㊱ 서킷브레이커 활성 시 신규진입 차단
                if circuit_active:
                    blocked_log.append({
                        "날짜": str(idx)[:10], "종목": name, "국면": curr_regime,
                        "Edge": round(adj_edge, 3),
                        "현재가": round(cp_now, 0), "Target_P": 0,
                        "슬리피지": slip, "차단이유": "㊱서킷브레이커(당일DD초과)"})
                    continue
                # ㉟ 멀티 타임프레임 필터
                if WEEKLY_TREND_REQUIRED:
                    wt_ok = calc_weekly_trend(res.loc[:idx])
                    if not wt_ok:
                        blocked_log.append({
                            "날짜": str(idx)[:10], "종목": name, "국면": curr_regime,
                            "Edge": round(adj_edge, 3),
                            "현재가": round(cp_now, 0), "Target_P": 0,
                            "슬리피지": slip, "차단이유": "㉟주봉추세불일치"})
                        continue
                # ㊴ 섹터 집중도 제한
                sec_ok, sec_reason = check_sector_concentration(
                    r_data["ticker"], positions, all_res_data)
                if not sec_ok:
                    sector_block_log.append({
                        "날짜": str(idx)[:10], "종목": name,
                        "사유": sec_reason})
                    blocked_log.append({
                        "날짜": str(idx)[:10], "종목": name, "국면": curr_regime,
                        "Edge": round(adj_edge, 3),
                        "현재가": round(cp_now, 0), "Target_P": 0,
                        "슬리피지": slip, "차단이유": sec_reason})
                    continue
                tp_price  = row.get("tp", cp_now * 1.01)
                filter_ok, filter_reason = check_slippage_filter(cp_now, tp_price, slip)

                if not filter_ok:
                    blocked_log.append({
                        "날짜": str(idx)[:10], "종목": name, "국면": curr_regime,
                        "Edge": round(row["edge"], 3),
                        "현재가": round(cp_now, 0), "Target_P": round(tp_price, 0),
                        "슬리피지": slip, "차단이유": filter_reason,
                    })
                    # ㉙ alt_pending 최대 5종목 제한 (연산 효율 + 방어 설계)
                    if len(alt_pending) < ALT_MAX_PENDING:
                        alt_pending[name] = {
                            "blocked_idx": day_i,
                            "kelly_mult":  ALT_KELLY_MULT,
                            "regime":      curr_regime,
                        }
                else:
                    kelly_info = calc_kelly_position(
                        dummy_win_rate, dummy_avg_win, dummy_avg_loss,
                        curr_regime, is_high_corr, mgr.cash)
                    # ㊲ 변동성 조절 사이징
                    vol_adj_amount = calc_vol_target_size(
                        res.loc[:idx], mgr.cash, kelly_info["amount"])
                    kelly_info["amount"] = vol_adj_amount
                    ok, allowed, cap_reason = mgr.can_buy(kelly_info["amount"], curr_regime)
                    if ok and allowed > 1000:
                        entry_rsn = (analyze_entry_reason(row, cp_map)
                                     + (f"[비중축소]" if cap_reason != "정상" else ""))
                        # [New-B] 당일 즉시 진입 → pending_entries 등록 후 T+1 시가 진입
                        # 단, next_open이 없는 마지막 데이터 날짜는 당일 종가로 폴백
                        if pd.notna(row.get("next_open", float("nan"))) and row.get("next_open", 0) > 0:
                            # 자본은 다음날 진입 시점에 차감하므로 오늘은 예약만 (can_buy는 이미 통과)
                            pending_entries[name] = {
                                "signal_row":   row,
                                "entry_reason": entry_rsn,
                                "allowed":      allowed,
                            }
                        else:
                            # 마지막 날 또는 next_open 누락: 당일 종가 폴백 (기존 동작 유지)
                            buy_p = apply_slippage(cp_now, "buy", slip)
                            mgr.on_buy(allowed)
                            pos.update({"amount": allowed, "buy_price": buy_p,
                                        "prev_price": buy_p, "hold_days": 0,
                                        "entry_row": idx, "entry_reason": entry_rsn,
                                        "peak_ret": 0.0})
                            pos["trailing"].reset()
                            trade_logs[name].append({
                                "종목": name, "진입일": str(idx)[:10], "청산일": None,
                                "보유기간": None, "국면": curr_regime,
                                "진입이유": entry_rsn + "[당일폴백]", "청산이유": "보유중",
                                "진입가": round(buy_p, 0), "청산가": None,
                                "수익률": None, "투자금": round(allowed, 0),
                                "손익금": None, "Kelly비율": round(allowed/INITIAL_CAPITAL, 4),
                                "ATR손절선": round(float(dynamic_sl), 4) if dynamic_sl is not None else None,
                                "㉔대안": "정규진입", "승패": "-",
                            })

        # ㉔ 대안 종목 재배치 처리
        for blocked_name, info in list(alt_pending.items()):
            if day_i < info["blocked_idx"] + ALT_WAIT_DAYS:
                continue  # 대기 기간 미충족
            # ㉙ Edge 내림차순 정렬 후 상위 ALT_MAX_CANDIDATES(10)개만 탐색
            candidates_alt = [
                r for r in all_res_data
                if r["name"] != blocked_name
                and positions[r["name"]]["amount"] == 0
                and idx in r["res"].index
            ]
            candidates_alt.sort(
                key=lambda r: r["res"].loc[idx, "edge"] if idx in r["res"].index else 0,
                reverse=True)
            candidates_alt = candidates_alt[:ALT_MAX_CANDIDATES]  # ㉙ 상위 10개 제한

            placed = False
            for alt_r in candidates_alt:
                alt_name    = alt_r["name"]
                alt_pos     = positions[alt_name]
                if alt_pos["amount"] > 0: continue
                if idx not in alt_r["res"].index: continue
                alt_row     = alt_r["res"].loc[idx]
                alt_regime  = curr_regime
                alt_thresh  = get_regime_threshold(alt_regime)
                if alt_row["edge"] < alt_thresh: continue

                # ㉕㉖ 누적 슬리피지 체크 (정규 필터 대신 ALT_FILTER_RATIO=3.5 적용)
                alt_slip    = get_slippage(alt_r["cluster"])
                alt_tp      = alt_row.get("tp", alt_row["종가"] * 1.01)
                wait_days_n = day_i - info["blocked_idx"]

                alt_ok, alt_reason, cost_detail = check_alt_cumulative_cost(
                    alt_row["종가"], alt_tp, alt_slip, wait_days_n,
                    regime=curr_regime)   # ㉚ 국면 전달
                if not alt_ok:
                    # 누적 비용 기록 (차단 사유 추적용)
                    alt_log.append({
                        "날짜":       str(idx)[:10],
                        "차단종목":   blocked_name,
                        "대안종목":   alt_name,
                        "결과":       "㉕누적비용차단",
                        "대안Edge":   round(alt_row["edge"], 3),
                        "투자금":     0,
                        "Kelly배율":  info["kelly_mult"],
                        "대기일":     wait_days_n,
                        "기회비용":   cost_detail.get("기회비용", 0),
                        "왕복슬리피지": cost_detail.get("왕복슬리피지", 0),
                        "총비용":     cost_detail.get("총비용", 0),
                        "기대수익":   cost_detail.get("기대수익", 0),
                        "필요최소수익": cost_detail.get("필요최소수익", 0),
                        "차단이유":   alt_reason,
                    })
                    continue

                # Kelly × ALT_KELLY_MULT
                alt_kelly   = calc_kelly_position(
                    dummy_win_rate, dummy_avg_win, dummy_avg_loss,
                    alt_regime, is_high_corr, mgr.cash,
                    kelly_mult=info["kelly_mult"])
                ok2, allowed2, _ = mgr.can_buy(alt_kelly["amount"], alt_regime)
                if not ok2 or allowed2 <= 1000: continue

                alt_sl      = alt_row.get("dynamic_sl", None)
                alt_atr     = abs(float(alt_sl)) if alt_sl is not None else 0.02
                buy_p2      = apply_slippage(alt_row["종가"], "buy", alt_slip)
                mgr.on_buy(allowed2)
                entry_rsn2  = (analyze_entry_reason(alt_row, alt_r["cp"])
                               + f"[㉔대안:{blocked_name}차단후재배치]")
                alt_pos.update({"amount": allowed2, "buy_price": buy_p2,
                                "prev_price": buy_p2, "hold_days": 0,
                                "entry_row": idx, "entry_reason": entry_rsn2,
                                "peak_ret": 0.0})
                alt_pos["trailing"].reset()
                trade_logs[alt_name].append({
                    "종목": alt_name, "진입일": str(idx)[:10], "청산일": None,
                    "보유기간": None, "국면": alt_regime,
                    "진입이유": entry_rsn2, "청산이유": "보유중",
                    "진입가": round(buy_p2, 0), "청산가": None,
                    "수익률": None, "투자금": round(allowed2, 0),
                    "손익금": None, "Kelly비율": round(allowed2/INITIAL_CAPITAL, 4),
                    "ATR손절선": round(float(alt_sl), 4) if alt_sl is not None else None,
                    "㉔대안": f"{blocked_name}차단후재배치", "승패": "-",
                })
                alt_log.append({
                    "날짜":       str(idx)[:10],
                    "차단종목":   blocked_name,
                    "대안종목":   alt_name,
                    "결과":       "배치완료",
                    "대안Edge":   round(alt_row["edge"], 3),
                    "투자금":     round(allowed2, 0),
                    "Kelly배율":  info["kelly_mult"],
                    "대기일":     day_i - info["blocked_idx"],
                    "기회비용":   cost_detail.get("기회비용", 0),
                    "왕복슬리피지": cost_detail.get("왕복슬리피지", 0),
                    "총비용":     cost_detail.get("총비용", 0),
                    "기대수익":   cost_detail.get("기대수익", 0),
                    "필요최소수익": cost_detail.get("필요최소수익", 0),
                    "차단이유":   "통과",
                })
                placed = True; break

            if placed or day_i > info["blocked_idx"] + 5:
                del alt_pending[blocked_name]  # 5일 이내 미배치 → 포기

        # ── ㉜ 주간 강제 청산 (금요일) + 자본 리셋 (월요일) ──
        if FRIDAY_FORCE_EXIT and is_friday(idx):  # [Issue-5 수정] is_friday()로 공휴일 제외
            for r_data in all_res_data:
                name2 = r_data["name"]
                pos2  = positions[name2]
                if pos2["amount"] <= 0: continue
                slip2   = get_slippage(r_data["cluster"])
                cp_fri  = r_data["res"]["종가"].get(idx, pos2["buy_price"])
                ret_fri = (cp_fri - pos2["buy_price"]) / pos2["buy_price"] \
                          if pos2["buy_price"] > 0 else 0

                # ── 수정된 원칙2: 보유 유지 판단 ──
                # 수익 중 + 트레일링 발동 중 OR AI점수 ≥ FRIDAY_HOLD_EDGE_THR → 이월
                edge_fri = r_data["res"].loc[idx, "edge"] \
                           if idx in r_data["res"].index else 0
                trail_active = pos2["trailing"].active \
                               if hasattr(pos2["trailing"], "active") else False
                # [Fix-1] RT hold_thr_eff(-0.03), OPT friday_hold_thr_eff(-0.03)와 동일 실효값
                hold_ok = (ret_fri > 0
                           and (trail_active or edge_fri >= FRIDAY_HOLD_EDGE_THR_EFF))

                if hold_ok:
                    # 이월 — 청산하지 않고 다음 주도 보유
                    _log_trade(trade_logs[name2], name2, pos2, idx, curr_regime,
                               "주간이월(보유유지)", cp_fri, cp_fri,
                               0, 0, None)   # 청산 없음 기록
                    continue

                # 청산
                sell_p2 = apply_slippage(cp_fri, "sell", slip2)
                pnl2    = pos2["amount"] * (sell_p2 - pos2["buy_price"]) / pos2["buy_price"]
                mgr.on_sell(pos2["amount"], pnl2)
                _log_trade(trade_logs[name2], name2, pos2, idx, curr_regime,
                           "주간청산(금요일)", cp_fri, sell_p2,
                           pos2["amount"], pnl2, None)
                pos2["amount"]    = 0
                pos2["hold_days"] = 0
                pos2["trailing"].reset()
                pos2["peak_ret"]  = 0.0

        # ── ㉜ 월요일 자본금 재계산 ──
        if WEEKLY_CAPITAL_RESET and idx.weekday() == 0:  # 월요일
            # 총자산 = 현금 + 평가금
            new_cap = mgr.total_capital + mgr.invested_amount
            # [Issue-1 수정] CAPITAL_FLOOR_RATIO 하방 보호 — RT monday_reset과 동일하게 적용
            prev_cap = week_start_capital  # 이번 주 시작 자본 (RT old_cap과 동일 역할)
            new_cap  = max(new_cap, prev_cap * CAPITAL_FLOOR_RATIO)
            # 이월 포지션 원투자금 합계 (이월 종목이 있으면 invested 재등록)
            # 이월 없을 때 0 → 기존 동작과 동일
            carry_inv = sum(
                pos2["amount"]
                for rd2 in all_res_data
                for pos2 in [positions[rd2["name"]]]
                if pos2["amount"] > 0
            )
            mgr.weekly_reset(new_cap, carry_invested=carry_inv)
            # [Issue-1 수정] 다음 주 하방 보호 기준 업데이트 (RT C["TOTAL_CAPITAL"] 갱신과 동일)
            week_start_capital = new_cap

        # 총 자산 기록
        unrealized = sum(
            pos["amount"] * (r_data["res"]["종가"].get(idx, pos["buy_price"])
                             / pos["buy_price"] - 1)
            for r_data in all_res_data
            for _, pos in [(r_data["name"], positions[r_data["name"]])]
            if pos["amount"] > 0 and pos["buy_price"] > 0
        )
        capital_hist.append(mgr.total_capital + mgr.invested_amount + unrealized)

    # 성과
    cap_ser      = pd.Series(capital_hist, index=all_dates)
    cum_strategy = cap_ser / INITIAL_CAPITAL
    mkt_ref      = all_res_data[0]["mkt"]["종가"].reindex(all_dates).ffill()
    cum_mkt      = mkt_ref / mkt_ref.iloc[0]
    total_ret    = cum_strategy.iloc[-1] - 1
    mkt_total    = cum_mkt.iloc[-1] - 1
    alpha        = total_ret - mkt_total
    rolling_max  = cum_strategy.cummax()
    mdd          = ((cum_strategy - rolling_max) / rolling_max).min()

    all_trades = []
    for logs in trade_logs.values(): all_trades.extend(logs)
    trades_df  = pd.DataFrame(all_trades)
    completed  = (trades_df[trades_df["청산이유"] != "보유중"]
                  if len(trades_df) else pd.DataFrame())
    win_rate   = ((completed["수익률"] > 0).mean()
                  if len(completed) and "수익률" in completed.columns else 0)

    # 주간 청산 + 이월 통계
    weekly_exits = completed[completed["청산이유"] == "주간청산(금요일)"] \
        if len(completed) and "청산이유" in completed.columns else pd.DataFrame()
    weekly_exit_count = len(weekly_exits)
    weekly_exit_wr    = ((weekly_exits["수익률"] > 0).mean()
                         if len(weekly_exits) else 0)
    carry_over_count  = len(trades_df[trades_df["청산이유"] == "주간이월(보유유지)"]) \
                        if len(trades_df) and "청산이유" in trades_df.columns else 0

    return {
        "total_ret": total_ret, "mkt_total": mkt_total, "alpha": alpha,
        "mdd": mdd, "win_rate": win_rate, "total_trades": len(completed),
        "final_capital": cap_ser.iloc[-1],
        "weekly_exit_count": weekly_exit_count,
        "weekly_exit_wr":    weekly_exit_wr,
        "carry_over_count":  carry_over_count,
        "cum_strategy": cum_strategy, "cum_mkt": cum_mkt,
        "trade_log": trades_df, "completed": completed,
        "blocked_log":    pd.DataFrame(blocked_log),
        "alt_log":        pd.DataFrame(alt_log),
        "trail_log":      pd.DataFrame(trail_log),
        "friction_log":   pd.DataFrame(friction_log),
        "rebalance_events": pd.DataFrame(rebalance_events)
            if rebalance_events else pd.DataFrame(),
        "time_stop_log": pd.DataFrame(time_stop_log),
        "circuit_log":   pd.DataFrame(circuit_log),
        "sector_block_log": pd.DataFrame(sector_block_log),
    }


def _log_trade(log_list, name, pos, idx, regime, reason,
               cp_now, sell_p, amount, pnl_amt, dynamic_sl):
    trade_r = (sell_p - pos["buy_price"]) / pos["buy_price"]
    log_list.append({
        "종목": name,
        "진입일": str(pos["entry_row"])[:10] if pos["entry_row"] else "",
        "청산일": str(idx)[:10], "보유기간": pos["hold_days"],
        "국면": regime, "진입이유": pos["entry_reason"],
        "청산이유": reason, "진입가": round(pos["buy_price"], 0),
        "청산가": round(cp_now, 0), "수익률": round(trade_r, 4),
        "투자금": round(amount, 0), "손익금": round(pnl_amt, 0),
        "Kelly비율": round(amount / INITIAL_CAPITAL, 4),
        "ATR손절선": round(float(dynamic_sl), 4) if dynamic_sl is not None else None,
        "㉔대안": "해당없음", "승패": "승" if trade_r > 0 else "패",
    })


# ══════════════════════════════════════════════════════
# 데이터 수집
# ══════════════════════════════════════════════════════

# yfinance 캐시 (코스피 지수 중복 조회 방지)
_YF_MKT_CACHE: dict = {}

def _yf_parse(raw: "pd.DataFrame") -> "pd.DataFrame":
    """yfinance MultiIndex 컬럼 → 단순 컬럼으로 정규화."""
    if isinstance(raw.columns, pd.MultiIndex):
        raw.columns = raw.columns.get_level_values(0)
    col_map = {"Close":"종가","Open":"시가","High":"고가","Low":"저가","Volume":"거래량"}
    return raw.rename(columns=col_map)

def _yf_get_mkt():
    """코스피 지수(^KS11) 캐시 조회."""
    if "mkt" not in _YF_MKT_CACHE:
        try:
            sd = pd.Timestamp(START_DATE).strftime("%Y-%m-%d")
            ed = pd.Timestamp(END_DATE).strftime("%Y-%m-%d")
            raw = yf.download("^KS11", start=sd, end=ed, auto_adjust=True, progress=False)
            raw = _yf_parse(raw)
            raw.index = pd.to_datetime(raw.index)
            _YF_MKT_CACHE["mkt"] = raw[["종가"]].copy()
        except Exception as e:
            print(f"  ⚠️  코스피 지수 조회 실패: {e}")
            _YF_MKT_CACHE["mkt"] = None
    return _YF_MKT_CACHE["mkt"]

def get_data(ticker):
    if YF_AVAILABLE:
        try:
            sd = pd.Timestamp(START_DATE).strftime("%Y-%m-%d")
            ed = pd.Timestamp(END_DATE).strftime("%Y-%m-%d")
            yt = f"{ticker}.KS"
            raw = yf.download(yt, start=sd, end=ed, auto_adjust=True, progress=False)
            if raw is None or len(raw) < 60:
                raise ValueError(f"데이터 부족: {len(raw) if raw is not None else 0}행")
            raw = _yf_parse(raw)
            raw.index = pd.to_datetime(raw.index)
            if "거래량" not in raw.columns: raw["거래량"] = 0
            if "시가"   not in raw.columns: raw["시가"]   = raw["종가"]
            if "고가"   not in raw.columns: raw["고가"]   = raw["종가"]
            if "저가"   not in raw.columns: raw["저가"]   = raw["종가"]
            raw["foreign_net"] = 0
            mkt = _yf_get_mkt()
            if mkt is None:
                raise ValueError("코스피 지수 없음")
            return raw, mkt, True
        except Exception as e:
            pass  # 폴백으로 진행
    np.random.seed(hash(ticker) % 2**32)
    dates = pd.bdate_range(START_DATE, END_DATE); n = len(dates)
    params = {"005930": (80000, 0.0003, 0.018), "005380": (200000, 0.0002, 0.020),
              "105560": (70000, 0.0002, 0.015), "031980": (50000, 0.0005, 0.030),
              "000660": (130000, 0.0003, 0.022), "035720": (60000, 0.0004, 0.028),
              "005490": (400000, 0.0002, 0.019), "373220": (400000, 0.0002, 0.024)}
    bp, dr, vl = params.get(ticker, (100000, 0.0002, 0.020))
    prices  = bp * np.exp(np.cumsum(np.random.normal(dr, vl, n)))
    volumes = np.random.lognormal(15, 0.5, n).astype(int)
    foreign = np.random.normal(0, volumes * 0.05).astype(int)
    # [New-B] 시가(open) 추가 — 다음날 시가 진입을 위해 필요
    # 실제 pykrx는 시가 포함. 모의 데이터는 전일 종가 대비 ±0.5% 갭으로 근사
    gap     = np.random.normal(0, 0.005, n)
    opens   = np.concatenate([[prices[0]], prices[:-1] * (1 + gap[1:])])
    df  = pd.DataFrame({"종가": prices, "시가": opens, "거래량": volumes, "foreign_net": foreign}, index=dates)
    mkt = pd.DataFrame({"종가": 2500 * np.exp(
        np.cumsum(np.random.normal(0.0002, 0.010, n)))}, index=dates)
    return df, mkt, False


# ══════════════════════════════════════════════════════
# 개별 종목 전처리
# ══════════════════════════════════════════════════════

def prepare_stock(name, ticker, use_kind_api=False):
    print(f"\n  {name} ({ticker}) 전처리 중...")
    cluster_name, cp = get_cluster_params(ticker)
    ci_t = stats.t.ppf(0.84, df=cp["T_DOF"])
    df, mkt, is_real = get_data(ticker)
    if len(df) < 60: return None

    regime_ser = calc_market_regime(mkt)
    mf   = calc_mf(df, cp["MF_CAP"]); tech = calc_tech(df, cp["TECH_CAP"])
    sigma_sel = calc_sigma(df); beta = calc_beta_eff(df, mkt)
    med_vol   = calc_med_vol(mkt); nc = df["종가"].shift(-1)
    # [New-B] 다음날 시가(next_open) — 신호 발생 당일이 아닌 다음 거래일 시가로 진입하기 위해 사용
    # pykrx OHLCV에는 시가 컬럼이 포함됨. 모의 데이터도 시가를 생성하도록 get_data 수정됨
    next_open = (df["시가"].shift(-1) if "시가" in df.columns
                 else df["종가"].shift(-1))  # 시가 없으면 다음날 종가로 대체 (안전 폴백)
    kind_s = build_kind_mom(ticker, df.index, use_api=use_kind_api)
    mom = (kind_s.reindex(df.index).fillna(0.5)
           if kind_s is not None else calc_mom_rsi(df))
    w_mf_e, w_tech_e, w_mom_e = calc_ensemble_weights(df, mf, tech, mom, nc)
    edge_e = mf*w_mf_e + tech*w_tech_e + mom*w_mom_e
    edge_s = mf*cp["W_MF"] + tech*cp["W_TECH"] + mom*cp["W_MOM"]
    edge   = edge_e.where(edge_e.notna(), edge_s).fillna(edge_s)
    mkt_vol   = mkt["종가"].pct_change().abs().rolling(20).mean().fillna(MED_VOL_FALLBACK)
    asym_k_ser = calc_dynamic_asym_k(mkt_vol, cp["ASYM_BASE"]).reindex(df.index).fillna(cp["ASYM_BASE"])
    is_anomaly, vol_zone_mult, avg_vol, vol_ratio = calc_vol_anomaly(df)
    tp, t_hi, t_lo, bias, sig_adj, d_mult = calc_targets(
        df, edge, sigma_sel, med_vol, beta, asym_k_ser, ci_t, vol_zone_mult)
    in_zone    = (nc >= t_lo) & (nc <= t_hi)
    dynamic_sl = calc_dynamic_stop_loss(df, cluster_name, regime_ser)

    res = pd.DataFrame({
        "종가": df["종가"], "거래량": df["거래량"],
        "mf": mf, "tech": tech, "mom": mom, "edge": edge, "bias": bias,
        "sigma_sel": sigma_sel, "sig_adj": sig_adj, "beta": beta,
        "med_vol": med_vol, "asym_k": asym_k_ser, "d_mult": d_mult,
        "vol_anomaly": is_anomaly, "vol_ratio": vol_ratio, "avg_vol": avg_vol,
        "tp": tp, "t_hi": t_hi, "t_lo": t_lo, "nc": nc, "in_zone": in_zone,
        "dynamic_sl": dynamic_sl,
        # [New-B] 다음날 시가 — 당일 신호 → 익일 시가 진입에 사용
        "next_open": next_open,
        # ㉝ RSI 다이버전스
        "rsi_divergence": calc_rsi_divergence(df),
    }).dropna()

    hr   = res["in_zone"].mean(); total = len(res)
    slip = get_slippage(cluster_name)
    sl_mean = res["dynamic_sl"].mean()
    sl_min  = res["dynamic_sl"].min()
    sl_max  = res["dynamic_sl"].max()

    # ㉑ 차단율 추정
    filter_blocked = sum(1 for _, row in res.iterrows()
                         if not check_slippage_filter(row["종가"], row["tp"], slip)[0])
    filter_block_rate = filter_blocked / len(res)

    slip_sens = calc_slippage_sensitivity(res, mkt, regime_ser, cluster_name, cp)
    bep       = slip_sens.get("bep")

    print(f"  클러스터: {cluster_name} | 슬리피지: {slip:.1%} | 적중률: {hr:.1%}")
    print(f"  ㉒ ATR 동적손절: 평균{sl_mean:.1%} [{sl_min:.1%}~{sl_max:.1%}]")
    print(f"  ㉓㉗ 트레일링 활성기준: +{TRAIL_ACTIVATE_RET:.0%} | "
          f"간격계수: ×{TRAIL_ATR_MULT.get(cluster_name, 2.0)} | "
          f"BEAR축소: ×{TRAIL_BEAR_MULT}")
    print(f"  ㉑ 차단율: {filter_block_rate:.1%} | ⑳ BEP: "
          f"{f'{bep:.1%}' if bep else '전구간손실'}")

    try: wf = adaptive_walk_forward(res, ci_t=ci_t, asym_base=cp["ASYM_BASE"])
    except: wf = None

    return {
        "name": name, "ticker": ticker, "cluster": cluster_name, "cp": cp,
        "hr": hr, "total": total, "hits": int(res["in_zone"].sum()),
        "upper_miss": (~(res["nc"] <= res["t_hi"])).mean(),
        "lower_miss": (~(res["nc"] >= res["t_lo"])).mean(),
        "res": res, "ci_t": ci_t, "wf": wf,
        "mkt": mkt, "regime_ser": regime_ser,
        "slip": slip, "slip_sens": slip_sens, "bep": bep,
        "sl_mean": sl_mean, "sl_min": sl_min, "sl_max": sl_max,
        "filter_block_rate": filter_block_rate,
    }


# ══════════════════════════════════════════════════════
# ⑯ 거래일지 패턴 분석
# ══════════════════════════════════════════════════════

def analyze_trade_journal(completed):
    if completed is None or len(completed) == 0: return
    df = completed[completed["수익률"].notna()].copy()
    if len(df) == 0: return
    print(f"\n{'='*62}")
    print("  ⑯ 거래 일지 패턴 분석")
    print(f"{'='*62}")
    rg = df.groupby("국면")["승패"].apply(lambda x: (x=="승").mean()).rename("승률").to_frame()
    rg["매매수"] = df.groupby("국면")["승패"].count()
    rg["승률"]   = rg["승률"].map("{:.1%}".format)
    print(f"\n  [국면별 승률]\n{rg.to_string()}")
    df["사유분류"] = df["청산이유"].apply(
        lambda x: "트레일링스탑" if "트레일링" in str(x)
                  else "ATR손절"  if "ATR손절"  in str(x)
                  else "익절"     if "익절"     in str(x) else "Edge하락")
    es = df.groupby("사유분류").agg(
        매매수=("수익률","count"),
        평균수익률=("수익률","mean"),
        승률=("승패", lambda x: (x=="승").mean()))
    es["평균수익률"] = es["평균수익률"].map("{:+.1%}".format)
    es["승률"]       = es["승률"].map("{:.1%}".format)
    print(f"\n  [청산 사유별 통계]\n{es.to_string()}")
    print(f"\n  평균 보유기간: {df['보유기간'].mean():.1f}일 | "
          f"전체 {len(df)}회 | 승률: {(df['승패']=='승').mean():.1%}")


# ══════════════════════════════════════════════════════
# 엑셀 리포트
# ══════════════════════════════════════════════════════

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
BAND_FILL   = PatternFill("solid", start_color="D6E4F0")
HIT_FILL    = PatternFill("solid", start_color="C6EFCE")
MISS_FILL   = PatternFill("solid", start_color="FFC7CE")
WARN_FILL   = PatternFill("solid", start_color="FFE699")
BULL_FILL   = PatternFill("solid", start_color="E2EFDA")
BEAR_FILL   = PatternFill("solid", start_color="FCE4D6")
WIN_FILL    = PatternFill("solid", start_color="A9D18E")
LOSE_FILL   = PatternFill("solid", start_color="FF7F7F")
TRAIL_FILL  = PatternFill("solid", start_color="BDD7EE")
ALT_FILL    = PatternFill("solid", start_color="E2CFFF")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BODY_FONT   = Font(name="Arial", size=9)
THIN   = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def _hdr(ws, row, col, val, width=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.border = BORDER
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if width: ws.column_dimensions[get_column_letter(col)].width = width
    return c

def _cell(ws, row, col, val, fmt=None, fill=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = BODY_FONT; c.border = BORDER
    c.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:  c.number_format = fmt
    if fill: c.fill = fill
    return c

def build_excel_report(all_res, port_pnl, risk,
                       filename="eqs_v1_1_report.xlsx"):
    wb = Workbook()

    # ── 시트1: 요약 ──
    ws = wb.active; ws.title = "요약 대시보드"
    ws.merge_cells("A1:O1")
    c = ws["A1"]; c.value = "EQS V1.2 백테스트 요약"
    c.font = Font(bold=True, size=14, color="1F4E79", name="Arial")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    hdrs = ["종목","클러스터","슬리피지","ATR손절평균","트레일기준",
            "㉑차단율","거래일","적중률","BEP","전략수익","코스피","알파","MDD","승률","최종자본"]
    wids = [16,10,7,9,8,8,7,7,8,9,8,8,7,7,12]
    for col,(h,w) in enumerate(zip(hdrs,wids),1): _hdr(ws,2,col,h,w)
    for ri,r in enumerate(all_res,3):
        fill = BAND_FILL if ri%2==0 else None
        bep_str = f"{r['bep']:.1%}" if r.get("bep") else "전구간손실"
        vals = [r["name"],r["cluster"],r["slip"],r["sl_mean"],
                f"+{TRAIL_ACTIVATE_RET:.0%}",r["filter_block_rate"],
                r["total"],r["hr"],bep_str,
                port_pnl["total_ret"],port_pnl["mkt_total"],port_pnl["alpha"],
                port_pnl["mdd"],port_pnl["win_rate"],port_pnl["final_capital"]]
        fmts = [None,None,"0.0%","0.0%",None,"0.0%","#,##0","0.0%",None,
                "0.0%","0.0%","0.0%","0.0%","0.0%","#,##0"]
        for col,(v,f) in enumerate(zip(vals,fmts),1): _cell(ws,ri,col,v,fmt=f,fill=fill)

    # 버전 이력
    sr = len(all_res)+5
    for i,line in enumerate([
        "v14~v16: 기본산식/RSI Mom/t-분포CI/비대칭CI",
        "v17~v18: ASYM_K동적/Walk-forward/KIND API/Cluster/Lag처리",
        "v19~v20: 앙상블/거래량이상/엑셀리포트/상관관계/국면/수익률시뮬/스케줄러",
        "v21: Kelly포지션/손절익절 트리거/거래일지",
        "v22: 클러스터별슬리피지/총비중한도/국면전환리밸런싱/BEP분석",
        "v23: ㉑슬리피지 진입금지 필터/㉒ATR 동적 손절선",
        "v24: ㉓ATR 트레일링 스탑/㉔대안 종목 자동 재배치",
        "v25: ㉕누적슬리피지 체크/㉖대안 전용 엄격 필터(3.5)",
        "v26: ㉗TrailingStop BEAR 가중/㉘기회비용 현실화(0.00014)/㉙후보군 효율+8종목 확대",
        "v1.1: ㊷코스피200 동적 유니버스 (200종목 자동 분류)",
    ]):
        ws.merge_cells(f"A{sr+i}:O{sr+i}")
        ws[f"A{sr+i}"].value = line
        ws[f"A{sr+i}"].font = Font(name="Arial", size=9,
                                   bold=(i==9), color=("1F4E79" if i==9 else "000000"))

    # ── 시트2: ㉓ 트레일링 스탑 로그 ──
    ws_t = wb.create_sheet("㉓㉗ 트레일링스탑")
    ws_t.merge_cells("A1:F1")
    ws_t["A1"].value = (f"㉓㉗ ATR 트레일링 스탑  "
                        f"(활성화: +{TRAIL_ACTIVATE_RET:.0%} / "
                        f"간격축소: +{TRAIL_TIGHTEN_RET:.0%}→×{TRAIL_TIGHTEN_MULT} / "
                        f"BEAR국면: ×{TRAIL_BEAR_MULT})")
    ws_t["A1"].font = Font(bold=True, size=12, color="1F4E79", name="Arial")
    ws_t.cell(3,1,"클러스터별 트레일링 ATR 계수").font=Font(bold=True,name="Arial",size=10)
    for col,h in enumerate(["클러스터","기본계수","BEAR국면(㉗)","고점+25%↑","BEAR+25%조합"],1):
        _hdr(ws_t,4,col,h,14)
    for ri,(cl,m) in enumerate(TRAIL_ATR_MULT.items(),5):
        bear_m   = round(m * TRAIL_BEAR_MULT, 2)
        tight_m  = round(m * TRAIL_TIGHTEN_MULT, 2)
        both_m   = round(m * TRAIL_BEAR_MULT * TRAIL_TIGHTEN_MULT, 2)
        fill = BEAR_FILL if ri % 2 == 0 else None
        _cell(ws_t,ri,1,cl,fill=fill);   _cell(ws_t,ri,2,f"×{m}",fill=fill)
        _cell(ws_t,ri,3,f"×{bear_m}",fill=BEAR_FILL)
        _cell(ws_t,ri,4,f"×{tight_m}",fill=fill)
        _cell(ws_t,ri,5,f"×{both_m}",fill=BEAR_FILL)

    # 트레일링 발동 이력
    ws_t.cell(10,1,"[트레일링 발동 이력]").font=Font(bold=True,name="Arial",size=10,color="1F4E79")
    for col,h in enumerate(["날짜","종목","최고수익률","청산수익률","수익보전"],1):
        _hdr(ws_t,11,col,h,14)
    trail_df = port_pnl.get("trail_log", pd.DataFrame())
    if len(trail_df) > 0:
        for ri,(_,row) in enumerate(trail_df.iterrows(),12):
            for col,v in enumerate([row.get("날짜",""),row.get("종목",""),
                                     row.get("최고수익률",""),row.get("청산수익률",""),
                                     row.get("트레일수익보전","")],1):
                _cell(ws_t,ri,col,v,fill=TRAIL_FILL)
    else:
        ws_t.cell(12,1,"트레일링 스탑 발동 이벤트 없음").font=Font(name="Arial",size=9,color="008000")

    # ── 시트3: ㉔㉕㉖ 대안 배치 로그 ──
    ws_a = wb.create_sheet("㉔~㉙ 대안배치")
    ws_a.merge_cells("A1:M1")
    ws_a["A1"].value = (f"㉔ 대안재배치  ㉕누적슬리피지  ㉖ALT_FILTER={ALT_FILTER_RATIO}  "
                        f"㉘기회비용={ALT_OPPORTUNITY_COST:.5f}/일(연{ALT_OPPORTUNITY_COST*250:.1%})  "
                        f"㉙상위{ALT_MAX_CANDIDATES}개/최대{ALT_MAX_PENDING}대기")
    ws_a["A1"].font = Font(bold=True, size=10, color="1F4E79", name="Arial")

    # 로직 설명
    desc_lines = [
        f"① ㉑ 필터 차단 발생 → 차단 종목 대기 큐 등록",
        f"② {ALT_WAIT_DAYS}거래일 대기 후 대안 종목 탐색 (Edge 높은 순)",
        f"③ ㉕ 누적비용 계산: 기회비용(0.01%×대기일) + 왕복슬리피지",
        f"④ 기대수익 ≥ 총비용 × {ALT_FILTER_RATIO} 이어야 통과 (정규 {SLIPPAGE_FILTER_RATIO}보다 엄격)",
        f"⑤ 통과 시 Kelly × {ALT_KELLY_MULT} 축소 투입 | 5거래일 내 미배치 → 포기",
    ]
    for i, line in enumerate(desc_lines, 2):
        ws_a.merge_cells(f"A{i}:M{i}")
        ws_a[f"A{i}"].value = line
        ws_a[f"A{i}"].font  = Font(name="Arial", size=9)

    # 비용 비교 테이블
    ws_a.cell(8, 1, "[ 정규 vs 대안 필터 비교 ]").font = Font(bold=True, name="Arial", size=10, color="1F4E79")
    for col, h in enumerate(["구분","필터비율","기회비용포함","설명"], 1):
        _hdr(ws_a, 9, col, h, 14)
    _cell(ws_a, 10, 1, "정규 진입");  _cell(ws_a, 10, 2, SLIPPAGE_FILTER_RATIO)
    _cell(ws_a, 10, 3, "미포함");     _cell(ws_a, 10, 4, "슬리피지×3.0만 체크")
    _cell(ws_a, 11, 1, "대안 배치",  fill=ALT_FILL)
    _cell(ws_a, 11, 2, ALT_FILTER_RATIO, fill=ALT_FILL)
    _cell(ws_a, 11, 3, "포함",        fill=ALT_FILL)
    _cell(ws_a, 11, 4, f"(기회비용+슬리피지)×{ALT_FILTER_RATIO}", fill=ALT_FILL)

    # 대안 배치 이력
    ws_a.cell(13, 1, "[ 대안 배치 이력 (통과 + 차단 모두 기록) ]").font = \
        Font(bold=True, name="Arial", size=10, color="1F4E79")
    alt_hdrs = ["날짜","차단종목","대안종목","결과","대안Edge",
                "기회비용","왕복슬리피지","총비용","기대수익","필요최소수익",
                "투자금","Kelly배율","차단이유"]
    alt_wids = [12,14,14,10,8,8,10,8,8,10,10,8,30]
    for col,(h,w) in enumerate(zip(alt_hdrs, alt_wids), 1):
        _hdr(ws_a, 14, col, h, w)

    alt_df = port_pnl.get("alt_log", pd.DataFrame())
    if len(alt_df) > 0:
        for ri, (_, row) in enumerate(alt_df.iterrows(), 15):
            is_placed = row.get("결과","") == "배치완료"
            fill = ALT_FILL if is_placed else WARN_FILL
            vals = [row.get("날짜",""),    row.get("차단종목",""),
                    row.get("대안종목",""), row.get("결과",""),
                    row.get("대안Edge",""),
                    row.get("기회비용", 0),   row.get("왕복슬리피지", 0),
                    row.get("총비용", 0),     row.get("기대수익", 0),
                    row.get("필요최소수익",0), row.get("투자금", 0),
                    row.get("Kelly배율",""),   row.get("차단이유","")]
            fmts = ["@",None,None,None,"0.000",
                    "0.000%","0.000%","0.000%","0.000%","0.000%",
                    "#,##0","0.0",None]
            for col,(v,f) in enumerate(zip(vals,fmts),1):
                _cell(ws_a, ri, col, v, fmt=f, fill=fill)

        # 요약 통계
        placed_n = len(alt_df[alt_df["결과"]=="배치완료"])
        blocked_n= len(alt_df[alt_df["결과"]=="㉕누적비용차단"])
        sr_a = len(alt_df) + 17
        ws_a.cell(sr_a, 1, f"배치완료: {placed_n}건  |  ㉕누적비용차단: {blocked_n}건  |  "
                            f"차단율: {blocked_n/(placed_n+blocked_n):.0%}" if (placed_n+blocked_n)>0 else "이벤트 없음")
        ws_a.cell(sr_a, 1).font = Font(bold=True, name="Arial", size=10, color="1F4E79")
    else:
        ws_a.cell(15, 1, "대안 배치 이벤트 없음").font = Font(name="Arial", size=9, color="008000")

    # ── 시트4: ㉑ 차단 로그 ──
    ws_b = wb.create_sheet("㉑ 진입금지필터")
    ws_b.merge_cells("A1:G1")
    ws_b["A1"].value = f"㉑ 슬리피지 진입금지 필터 로그 (기대수익 ≥ 왕복슬리피지×{SLIPPAGE_FILTER_RATIO})"
    ws_b["A1"].font  = Font(bold=True, size=12, color="1F4E79", name="Arial")
    for col,h in enumerate(["날짜","종목","국면","Edge","현재가","Target_P","차단이유"],1):
        _hdr(ws_b,2,col,h,14)
    blocked = port_pnl.get("blocked_log", pd.DataFrame())
    if len(blocked) > 0:
        for ri,(_,row) in enumerate(blocked.iterrows(),3):
            vals = [row.get("날짜",""),row.get("종목",""),row.get("국면",""),
                    row.get("Edge",""),row.get("현재가",""),
                    row.get("Target_P",""),row.get("차단이유","")]
            fmts = ["@",None,None,"0.000","#,##0","#,##0",None]
            for col,(v,f) in enumerate(zip(vals,fmts),1):
                _cell(ws_b,ri,col,v,fmt=f,fill=WARN_FILL)
    else:
        ws_b.cell(3,1,"차단 없음").font=Font(name="Arial",size=9,color="008000")

    # ── 시트5: 거래 일지 ──
    ws_j = wb.create_sheet("⑯ 거래일지")
    ws_j.merge_cells("A1:P1")
    ws_j["A1"].value = "⑯ 거래 일지 (㉓트레일링+㉔대안재배치 반영)"
    ws_j["A1"].font  = Font(bold=True, size=12, color="1F4E79", name="Arial")
    jh = ["종목","진입일","청산일","보유기간","국면","진입이유","청산이유",
          "진입가","청산가","수익률","투자금","손익금","Kelly비율","ATR손절선","㉔대안","승패"]
    jw = [12,12,12,6,5,26,22,9,9,7,9,9,7,8,14,5]
    for col,(h,w) in enumerate(zip(jh,jw),1): _hdr(ws_j,2,col,h,w)
    completed = port_pnl.get("completed", pd.DataFrame())
    if len(completed) > 0:
        for ri,(_,row) in enumerate(completed.iterrows(),3):
            win  = row.get("승패", "-")
            rsn  = str(row.get("청산이유",""))
            fill = (TRAIL_FILL if "트레일링" in rsn
                    else WIN_FILL  if win=="승"
                    else LOSE_FILL if win=="패"
                    else BEAR_FILL if row.get("국면","")=="BEAR"
                    else BULL_FILL if row.get("국면","")=="BULL" else None)
            vals = [row.get("종목",""),
                    str(row.get("진입일",""))[:10],
                    str(row.get("청산일",""))[:10] if row.get("청산일") else "",
                    row.get("보유기간",""),row.get("국면",""),
                    row.get("진입이유",""),row.get("청산이유",""),
                    row.get("진입가",""),row.get("청산가",""),
                    row.get("수익률",""),row.get("투자금",""),
                    row.get("손익금",""),row.get("Kelly비율",""),
                    row.get("ATR손절선",""),row.get("㉔대안",""),win]
            fmts = [None,"@","@","#,##0",None,None,None,
                    "#,##0","#,##0","0.0%","#,##0","#,##0","0.0%","0.0%",None,None]
            for col,(v,f) in enumerate(zip(vals,fmts),1):
                _cell(ws_j,ri,col,v,fmt=f,fill=fill)

    # ── 시트5.5: ㉛ 교체 마찰 비용 로그 ──
    ws_fr = wb.create_sheet("㉛ 교체마찰비용")
    ws_fr.merge_cells("A1:I1")
    ws_fr["A1"].value = (f"㉛ 교체 마찰 비용 로그  "
                         f"(매도마찰×{HOLD_FRICTION_MULT} / Edge우위>{HOLD_FRICTION_EDGE_GAP_MULT}×교체총마찰)")
    ws_fr["A1"].font = Font(bold=True, size=11, color="1F4E79", name="Arial")

    fr_desc = [
        "  보유 종목은 이미 진입 슬리피지를 지불한 상태 → 교체 시 추가 비용 발생",
        f"  교체 총마찰 = 현종목 매도슬리피지 × {HOLD_FRICTION_MULT}(가중) + 대안 매수슬리피지",
        f"  허용 조건: 대안Edge - 현보유Edge > 교체총마찰 × {HOLD_FRICTION_EDGE_GAP_MULT}",
        "  효과: 미미한 Edge 차이로 인한 잦은 교체 방지 / 모든 종목에 동일 적용(백테스트 유효성 유지)",
    ]
    for i, line in enumerate(fr_desc, 2):
        ws_fr.merge_cells(f"A{i}:I{i}")
        ws_fr[f"A{i}"].value = line
        ws_fr[f"A{i}"].font  = Font(name="Arial", size=9)

    fr_hdrs = ["날짜","보유종목","대안종목","결과","보유Edge","대안Edge",
               "Edge우위","필요우위","교체총마찰"]
    fr_wids = [12,14,14,14,8,8,8,8,10]
    for col,(h,w) in enumerate(zip(fr_hdrs,fr_wids),1):
        _hdr(ws_fr,7,col,h,w)

    friction_df2 = port_pnl.get("friction_log", pd.DataFrame())
    if len(friction_df2) > 0:
        for ri,(_,row) in enumerate(friction_df2.iterrows(),8):
            is_hold = row.get("결과","") == "교체보류(마찰비용)"
            fill    = WARN_FILL if is_hold else WIN_FILL
            vals = [row.get("날짜",""),    row.get("보유종목",""),
                    row.get("대안종목",""), row.get("결과",""),
                    row.get("보유Edge",0),  row.get("대안Edge",0),
                    row.get("Edge우위",0),  row.get("필요우위",0),
                    row.get("교체총마찰",0)]
            fmts = ["@",None,None,None,
                    "0.000","0.000","0.000","0.000","0.000%"]
            for col,(v,f) in enumerate(zip(vals,fmts),1):
                _cell(ws_fr,ri,col,v,fmt=f,fill=fill)

        hold_n2    = len(friction_df2[friction_df2["결과"]=="교체보류(마찰비용)"])
        replace_n2 = len(friction_df2[friction_df2["결과"]=="교체허용"])
        sr_fr = len(friction_df2) + 10
        ws_fr.merge_cells(f"A{sr_fr}:I{sr_fr}")
        ws_fr[f"A{sr_fr}"].value = (f"교체보류: {hold_n2}건  |  교체허용: {replace_n2}건  |  "
                                     f"보류율: {hold_n2/(hold_n2+replace_n2):.0%}"
                                     if (hold_n2+replace_n2)>0 else "이벤트 없음")
        ws_fr[f"A{sr_fr}"].font = Font(bold=True, name="Arial", size=10, color="1F4E79")
    else:
        ws_fr.cell(8,1,"교체 마찰 이벤트 없음").font = Font(name="Arial",size=9,color="008000")

    # ── 시트6~9: 종목별 일별 ──
    for r in all_res:
        ws_d = wb.create_sheet(r["name"][:20])
        cols = ["날짜","종가","Edge","국면","ATR손절선","트레일기준","Target_P","하단","상단","실제종가","적중"]
        wids2= [12,10,7,6,9,9,10,10,10,10,6]
        for col,(h,w) in enumerate(zip(cols,wids2),1): _hdr(ws_d,1,col,h,w)
        ra = r["regime_ser"].reindex(r["res"].index).ffill().fillna("SIDE")
        for ri,( idx,row) in enumerate(r["res"].tail(100).iterrows(),2):
            rg2 = ra.get(idx,"SIDE"); hit = row.get("in_zone",False)
            fill= (BULL_FILL if rg2=="BULL" else BEAR_FILL if rg2=="BEAR"
                   else HIT_FILL if hit else MISS_FILL)
            trail_price_ref = row["종가"] * (1 + TRAIL_ACTIVATE_RET)
            vals = [idx.strftime("%Y-%m-%d"),row["종가"],row["edge"],rg2,
                    row.get("dynamic_sl",STOP_LOSS_FALLBACK),
                    f"+{TRAIL_ACTIVATE_RET:.0%}활성",
                    row["tp"],row["t_lo"],row["t_hi"],row["nc"],
                    "✅" if hit else "❌"]
            fmts = ["@","#,##0","0.000",None,"0.0%",None,
                    "#,##0","#,##0","#,##0","#,##0",None]
            for col,(v,f) in enumerate(zip(vals,fmts),1): _cell(ws_d,ri,col,v,fmt=f,fill=fill)

    wb.save(filename)
    print(f"  📊 엑셀 리포트 저장: {filename}")
    return filename


# ══════════════════════════════════════════════════════
# 시각화
# ══════════════════════════════════════════════════════

def plot_results(all_res, port_pnl):
    n   = len(all_res)
    fig = plt.figure(figsize=(14, n*4+9))
    gs  = fig.add_gridspec(n+2, 2, hspace=0.45, wspace=0.3)
    fig.suptitle("EQS V1.2 Backtest", fontsize=13, fontweight="bold")

    for i, r in enumerate(all_res):
        ax   = fig.add_subplot(gs[i,:])
        df_r = r["res"].tail(60)
        ax.fill_between(df_r.index,df_r["tp"],df_r["t_hi"],alpha=0.15,color="blue",label="상단CI")
        ax.fill_between(df_r.index,df_r["t_lo"],df_r["tp"],alpha=0.15,color="orange",label="하단CI")
        ax.plot(df_r.index,df_r["종가"],color="black",lw=1.2,label="실제종가")
        ax.plot(df_r.index,df_r["tp"],color="red",lw=1,ls="--",label="Target_P")
        # ㉒ ATR 손절선
        if "dynamic_sl" in df_r.columns:
            ax.plot(df_r.index, df_r["종가"]*(1+df_r["dynamic_sl"]),
                    color="crimson", lw=0.7, ls=":", alpha=0.6, label="ATR손절선")
        # ㉓ 트레일링 활성 기준선
        ax.axhline(df_r["종가"].iloc[-1] * (1 + TRAIL_ACTIVATE_RET),
                   color="steelblue", lw=0.5, ls="--", alpha=0.4)
        hit  = df_r[df_r["in_zone"]==True]; miss = df_r[df_r["in_zone"]==False]
        ax.scatter(hit.index,  hit["nc"],  color="blue",   s=12, zorder=5)
        ax.scatter(miss.index, miss["nc"], color="red",    s=12, zorder=5)
        anom = df_r[df_r["vol_anomaly"]==True]
        ax.scatter(anom.index, anom["종가"], color="purple", marker="^", s=50, zorder=6)
        ax.set_title(
            f"{r['name']} [{r['cluster']}] | 적중률 {r['hr']:.1%} | "
            f"슬리피지 {r['slip']:.1%} | 트레일+{TRAIL_ACTIVATE_RET:.0%}활성 | "
            f"차단{r['filter_block_rate']:.1%}", fontsize=9)
        ax.legend(fontsize=6, loc="upper left")
        ax.tick_params(axis="x", rotation=30, labelsize=7)
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f"{x:,.0f}"))

    # 포트폴리오 누적수익률
    ax_pnl = fig.add_subplot(gs[n,:])
    ax_pnl.plot(port_pnl["cum_strategy"].index, port_pnl["cum_strategy"],
                color="navy", lw=2, label="포트폴리오 전략")
    ax_pnl.plot(port_pnl["cum_mkt"].index, port_pnl["cum_mkt"],
                color="gray", lw=1, ls="--", label="코스피 B&H")
    ax_pnl.axhline(1.0, color="gray", lw=0.7, ls=":")
    ax_pnl.set_title(
        f"포트폴리오 누적수익률 | "
        f"전략 {port_pnl['total_ret']:+.1%} / "
        f"코스피 {port_pnl['mkt_total']:+.1%} / "
        f"알파 {port_pnl['alpha']:+.1%} | "
        f"트레일발동 {len(port_pnl['trail_log'])}건 / "
        f"대안배치 {len(port_pnl['alt_log'])}건", fontsize=9)
    ax_pnl.legend(fontsize=8)
    ax_pnl.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f"{x:.2f}x"))
    ax_pnl.tick_params(axis="x", rotation=30, labelsize=7)

    # ㉓㉔ 이벤트 수 비교 바 차트
    ax_ev = fig.add_subplot(gs[n+1,:])
    labels = [r["name"] for r in all_res]
    trail_counts = [
        len([t for t in port_pnl["trail_log"].get("종목", pd.Series()).tolist()
             if t == r["name"]]) for r in all_res
    ] if len(port_pnl["trail_log"]) > 0 else [0]*len(all_res)
    blocked_counts = [
        len(port_pnl["blocked_log"][port_pnl["blocked_log"]["종목"]==r["name"]])
        if len(port_pnl["blocked_log"])>0 else 0
        for r in all_res
    ]
    alt_counts = [
        len(port_pnl["alt_log"][port_pnl["alt_log"]["대안종목"]==r["name"]])
        if len(port_pnl["alt_log"])>0 else 0
        for r in all_res
    ]
    x = np.arange(len(labels)); w = 0.25
    ax_ev.bar(x-w, blocked_counts, w, label="㉑차단",   color="#F44336", alpha=0.8)
    ax_ev.bar(x,   trail_counts,   w, label="㉓트레일",  color="#2196F3", alpha=0.8)
    ax_ev.bar(x+w, alt_counts,     w, label="㉔대안배치", color="#9C27B0", alpha=0.8)
    ax_ev.set_xticks(x); ax_ev.set_xticklabels(labels, fontsize=9)
    ax_ev.set_ylabel("이벤트 수")
    ax_ev.set_title("㉑㉓㉔ 이벤트 발생 현황", fontsize=10)
    ax_ev.legend(fontsize=8)

    plt.savefig("eqs_v1_2_backtest_result.png", dpi=150, bbox_inches="tight")
    print("  📊 차트 저장: eqs_v1_2_backtest_result.png")
    plt.close()


# ══════════════════════════════════════════════════════
# ⑬ 스케줄러 + 이메일
# ══════════════════════════════════════════════════════

def send_email_report(all_res, port_pnl, excel_file):
    try:
        msg = MIMEMultipart()
        msg["From"] = EMAIL_CONFIG["sender"]; msg["To"] = EMAIL_CONFIG["receiver"]
        msg["Subject"] = f"EQS V1.2 리포트 ({END_DATE})"
        lines = [
            f"포트폴리오: {port_pnl['total_ret']:+.1%} | 알파: {port_pnl['alpha']:+.1%}",
            f"㉓ 트레일링발동: {len(port_pnl['trail_log'])}건 | "
            f"㉔ 대안배치: {len(port_pnl['alt_log'])}건\n",
        ]
        for r in all_res:
            lines.append(f"[{r['name']}] ATR손절:{r['sl_mean']:.1%} | "
                         f"차단:{r['filter_block_rate']:.1%}")
        msg.attach(MIMEText("\n".join(lines), "plain", "utf-8"))
        if os.path.exists(excel_file):
            with open(excel_file,"rb") as f:
                part = MIMEBase("application","octet-stream"); part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition",
                            f'attachment; filename="{os.path.basename(excel_file)}"')
            msg.attach(part)
        with smtplib.SMTP(EMAIL_CONFIG["smtp"], EMAIL_CONFIG["port"]) as server:
            server.starttls(); server.login(EMAIL_CONFIG["sender"], EMAIL_CONFIG["password"])
            server.send_message(msg)
        print("  ✉️  이메일 발송 완료")
    except Exception as e:
        print(f"  ⚠️  이메일 발송 실패: {e}")

def scheduled_job():
    print(f"\n  🕐 자동 실행: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}")
    all_res = []
    for name, ticker in TICKERS.items():
        r = prepare_stock(name, ticker, use_kind_api=True)
        if r: all_res.append(r)
    if not all_res: return
    risk     = calc_portfolio_risk(all_res)
    port_pnl = calc_portfolio_pnl(all_res, all_res[0]["regime_ser"], risk["is_high_corr"])
    fname    = f"eqs_v1_2_{pd.Timestamp.now().strftime('%Y%m%d')}.xlsx"
    ef       = build_excel_report(all_res, port_pnl, risk, fname)
    send_email_report(all_res, port_pnl, ef)

def run_scheduler():
    if not SCHEDULE_AVAILABLE: print("⚠️  schedule 미설치"); return
    print("  ⑬ 스케줄러 시작 (매일 15:40) / 종료: Ctrl+C\n")
    schedule.every().day.at("15:40").do(scheduled_job)
    while True: schedule.run_pending(); time.sleep(30)


# ══════════════════════════════════════════════════════
# 메인
# ══════════════════════════════════════════════════════

def main():
    if "--scheduler" in sys.argv: run_scheduler(); return

    print("\n" + "="*62)
    print("  📊 EQS V1.2 백테스트 (Edge Quant Signal) (코스피200 동적 유니버스)")
    print("="*62)
    print(f"  기간: {START_DATE} ~ {END_DATE}")
    print(f"  ①~㉙ v26.0 전 기능 유지")
    print(f"  ㉚ ALT_FILTER 국면 연동  "
          f"BULL={ALT_FILTER_BY_REGIME['BULL']} / "
          f"SIDE={ALT_FILTER_BY_REGIME['SIDE']} / "
          f"BEAR={ALT_FILTER_BY_REGIME['BEAR']}")
    print(f"  ㉛ 교체 마찰 비용  "
          f"(매도×{HOLD_FRICTION_MULT} + 매수 / Edge우위>{HOLD_FRICTION_EDGE_GAP_MULT}배)")

    USE_KIND_API = False
    print(f"  KIND API: {'사용' if USE_KIND_API else '미사용(RSI fallback)'}\n")

    all_res = []
    for name, ticker in TICKERS.items():
        r = prepare_stock(name, ticker, use_kind_api=USE_KIND_API)
        if r: all_res.append(r)
    if not all_res: return

    risk = calc_portfolio_risk(all_res)

    print(f"\n{'='*62}")
    print("  포트폴리오 통합 시뮬레이션 중 (㉑~㉛ 반영)...")
    print(f"{'='*62}")
    port_pnl = calc_portfolio_pnl(all_res, all_res[0]["regime_ser"], risk["is_high_corr"])

    print(f"\n  [포트폴리오 성과]")
    print(f"  초기: {INITIAL_CAPITAL:,}원 → 최종: {port_pnl['final_capital']:,.0f}원")
    print(f"  전략: {port_pnl['total_ret']:+.1%} | 코스피: {port_pnl['mkt_total']:+.1%} | "
          f"알파: {port_pnl['alpha']:+.1%}")
    print(f"  MDD: {port_pnl['mdd']:.1%} | 승률: {port_pnl['win_rate']:.1%} | "
          f"매매: {port_pnl['total_trades']}회")
    print(f"  ㉜ 주간청산: {port_pnl['weekly_exit_count']}회 "
          f"| 주간청산 승률: {port_pnl['weekly_exit_wr']:.1%} "
          f"| 이월보유: {port_pnl['carry_over_count']}회")
    friction_df  = port_pnl.get("friction_log", pd.DataFrame())
    hold_n    = len(friction_df[friction_df["결과"]=="교체보류(마찰비용)"]) if len(friction_df) else 0
    replace_n = len(friction_df[friction_df["결과"]=="교체허용"])            if len(friction_df) else 0
    print(f"  ㉓ 트레일링 발동: {len(port_pnl['trail_log'])}건 | "
          f"㉔ 대안 시도: {len(port_pnl['alt_log'])}건 | "
          f"㉑ 차단: {len(port_pnl['blocked_log'])}건")
    alt_placed  = len(port_pnl['alt_log'][port_pnl['alt_log']['결과']=='배치완료']) \
                  if len(port_pnl['alt_log'])>0 else 0
    alt_blocked = len(port_pnl['alt_log'][port_pnl['alt_log']['결과']=='㉕누적비용차단']) \
                  if len(port_pnl['alt_log'])>0 else 0
    print(f"  ㉕㉖㉚ 대안배치완료: {alt_placed}건 | 누적비용차단: {alt_blocked}건")
    print(f"  ㉛ 교체마찰 보류: {hold_n}건 | 교체허용: {replace_n}건")
    ts_log = port_pnl.get("time_stop_log", pd.DataFrame())
    cb_log = port_pnl.get("circuit_log", pd.DataFrame())
    sb_log = port_pnl.get("sector_block_log", pd.DataFrame())
    print(f"  ㊳ 타임스탑: {len(ts_log)}건 | ㊱ 서킷브레이커: {len(cb_log)}건 | "
          f"㊴ 섹터집중차단: {len(sb_log)}건")

    analyze_trade_journal(port_pnl.get("completed"))

    print(f"\n{'='*62}")
    print("  📋 종목별 요약")
    print(f"{'='*62}")
    print(f"  {'종목':<16} {'클러스터':<10} {'ATR손절':>7}  "
          f"{'트레일간격':>8}  {'㉑차단율':>7}  {'BEP':>8}")
    for r in all_res:
        tm = TRAIL_ATR_MULT.get(r["cluster"], 2.0)
        bep_str = f"{r['bep']:.1%}" if r.get("bep") else "N/A"
        print(f"  {r['name']:<16} {r['cluster']:<10} "
              f"{r['sl_mean']:>7.1%}  ATR×{tm:<5}  "
              f"{r['filter_block_rate']:>7.1%}  {bep_str:>8}")

    plot_results(all_res, port_pnl)

    print(f"\n{'='*62}")
    print("  엑셀 리포트 생성 중...")
    build_excel_report(all_res, port_pnl, risk, "eqs_v1_1_report.xlsx")

    print(f"\n{'='*62}")
    print("  ✅ EQS V1.2 백테스트 완료")
    print(f"{'='*62}")
    print("\n  [실전 전환 체크리스트]")
    print("  □ pip install yfinance             → 실제 KRX 데이터")
    print(f"  □ ALT_FILTER_BY_REGIME 조정        → BULL={ALT_FILTER_BY_REGIME['BULL']} / SIDE={ALT_FILTER_BY_REGIME['SIDE']} / BEAR={ALT_FILTER_BY_REGIME['BEAR']}")
    print(f"  □ HOLD_FRICTION_MULT 조정          → 현재 {HOLD_FRICTION_MULT} (높일수록 교체 억제)")
    print(f"  □ HOLD_FRICTION_EDGE_GAP_MULT 조정 → 현재 {HOLD_FRICTION_EDGE_GAP_MULT}")
    print(f"  □ ALT_OPPORTUNITY_COST 조정        → 현재 {ALT_OPPORTUNITY_COST:.5f}/일 = 연{ALT_OPPORTUNITY_COST*250:.1%}")
    print(f"  □ 유니버스                          → {len(TICKERS)}종목 (코스피200)")
    print("  □ USE_KIND_API = True              → 공시 감성 연동")
    print("  □ --scheduler                      → 자동 실행")


if __name__ == "__main__":
    main()